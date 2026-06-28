

import os
from threading import Thread
from http.server import HTTPServer, BaseHTTPRequestHandler

# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                    IMPORTS & CONFIG                          в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ
import httpx
import sqlite3
import asyncio
import io
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ContextTypes, MessageHandler, filters

TOKEN = os.environ.get("TOKEN", "")
ADMIN_ID = int(os.environ.get("ADMIN_ID", "0"))
GOLD_API_KEY = "goldapi-d23da414dfdcbbe06a2e2ce8d28a095c-io"


# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                       DATABASE                               в•‘
# в•‘  ШӘЩҲШ§ШЁШ№ ШҜЫҢШӘШ§ШЁЫҢШі - ШҜШіШӘ ЩҶШІЩҶ                                    в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ
def init_db():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        user_id INTEGER PRIMARY KEY, username TEXT, first_name TEXT,
        status TEXT DEFAULT 'pending', approved_at TEXT,
        expires_at TEXT, daily_msg INTEGER DEFAULT 1)''')
    c.execute('''CREATE TABLE IF NOT EXISTS watchlist (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER, symbol TEXT, market TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS alarms (
        id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER,
        symbol TEXT, market TEXT, target REAL, direction TEXT,
        active INTEGER DEFAULT 1, created_at TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS user_state (
        user_id INTEGER PRIMARY KEY, state TEXT, data TEXT)''')
    try:
        c.execute("ALTER TABLE users ADD COLUMN daily_msg INTEGER DEFAULT 1")
    except:
        pass
    conn.commit()
    conn.close()

def get_user(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row

def save_user(user_id, username, first_name):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT OR IGNORE INTO users (user_id, username, first_name, status, daily_msg) VALUES (?, ?, ?, 'pending', 1)",
              (user_id, username, first_name))
    conn.commit()
    conn.close()

def approve_user(user_id):
    now = datetime.now()
    expires = now + timedelta(days=365)
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE users SET status='active', approved_at=?, expires_at=?, daily_msg=1 WHERE user_id=?",
              (now.strftime("%Y-%m-%d %H:%M"), expires.strftime("%Y-%m-%d %H:%M"), user_id))
    conn.commit()
    conn.close()

def reject_user(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE users SET status='rejected' WHERE user_id=?", (user_id,))
    conn.commit()
    conn.close()

def set_state(user_id, state, data=""):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO user_state (user_id, state, data) VALUES (?, ?, ?)", (user_id, state, data))
    conn.commit()
    conn.close()

def get_state(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT state, data FROM user_state WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row if row else (None, None)

def clear_state(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("DELETE FROM user_state WHERE user_id=?", (user_id,))
    conn.commit()
    conn.close()

def get_watchlist(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id, symbol, market FROM watchlist WHERE user_id=?", (user_id,))
    rows = c.fetchall()
    conn.close()
    return rows

def add_watchlist(user_id, symbol, market):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id FROM watchlist WHERE user_id=? AND symbol=?", (user_id, symbol))
    if c.fetchone():
        conn.close()
        return False
    c.execute("INSERT INTO watchlist (user_id, symbol, market) VALUES (?, ?, ?)", (user_id, symbol, market))
    conn.commit()
    conn.close()
    return True

def remove_watchlist(wid):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("DELETE FROM watchlist WHERE id=?", (wid,))
    conn.commit()
    conn.close()

def get_alarms(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id, symbol, market, target, direction FROM alarms WHERE user_id=? AND active=1", (user_id,))
    rows = c.fetchall()
    conn.close()
    return rows

def add_alarm(user_id, symbol, market, target, direction):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO alarms (user_id, symbol, market, target, direction, active, created_at) VALUES (?, ?, ?, ?, ?, 1, ?)",
              (user_id, symbol, market, target, direction, datetime.now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def deactivate_alarm(alarm_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE alarms SET active=0 WHERE id=?", (alarm_id,))
    conn.commit()
    conn.close()

def remove_alarm(alarm_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("DELETE FROM alarms WHERE id=?", (alarm_id,))
    conn.commit()
    conn.close()

def get_all_active_alarms():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id, user_id, symbol, market, target, direction FROM alarms WHERE active=1")
    rows = c.fetchall()
    conn.close()
    return rows


# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                  MODULE: FINANCIAL                           в•‘
# в•‘                                                              в•‘
# в•‘  ШЁШұШ§ЫҢ ШӯШ°ЩҒ Ъ©Ш§Щ…Щ„ Ш§ЫҢЩҶ Щ…Ш§ЪҳЩҲЩ„:                                   в•‘
# в•‘  Ыұ. Ш§ЫҢЩҶ ШЁЩ„ЩҲЪ© ШұЩҲ ЩҫШ§Ъ© Ъ©ЩҶ (Ш§ШІ Ш§ЫҢЩҶШ¬Ш§ ШӘШ§ Ш®Ш· в•җв•җв•җ ШЁШ№ШҜЫҢ)          в•‘
# в•‘  ЫІ. ШҜШұ MAIN MENU Ш®Ш· financial ШұЩҲ Ъ©Ш§Щ…ЩҶШӘ Ъ©ЩҶ                  в•‘
# в•‘  Ыі. ШҜШұ ROUTER Ш®Ш· financial ШұЩҲ Ъ©Ш§Щ…ЩҶШӘ Ъ©ЩҶ                     в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ

IRAN_SYMBOLS = {
    "USD": "price_dollar_rl", "EUR": "price_eur", "GBP": "price_gbp",
    "GOLD18": "geram18", "COIN": "sekee", "HALFCOIN": "nim",
    "QTRCOIN": "rob", "SILVER": "silver_999", "MITHGAL": "mesghal",
}
CRYPTO_SYMBOLS = {
    "BTC": "bitcoin", "ETH": "ethereum", "USDT": "tether",
    "BNB": "binancecoin", "TRX": "tron", "SOL": "solana",
    "ADA": "cardano", "XRP": "ripple", "DOGE": "dogecoin",
    "DOT": "polkadot", "MATIC": "matic-network", "LTC": "litecoin",
    "AVAX": "avalanche-2", "LINK": "chainlink", "UNI": "uniswap",
}
GOLDAPI_SYMBOLS = {
    "GOLD": ("XAU", "USD"), "XAUUSD": ("XAU", "USD"),
    "SILVER": ("XAG", "USD"), "XAGUSD": ("XAG", "USD"),
}
FOREX_YAHOO = {
    "EURUSD": "EURUSD%3DX", "GBPUSD": "GBPUSD%3DX",
    "USDCAD": "USDCAD%3DX", "USDJPY": "USDJPY%3DX",
    "AUDUSD": "AUDUSD%3DX", "USDCHF": "USDCHF%3DX",
    "NZDUSD": "NZDUSD%3DX", "EURGBP": "EURGBP%3DX",
    "EURJPY": "EURJPY%3DX", "GBPJPY": "GBPJPY%3DX",
    "OIL": "BZ%3DF", "BRENT": "BZ%3DF", "WTI": "CL%3DF",
}

def detect_market(symbol):
    s = symbol.upper()
    if s in IRAN_SYMBOLS: return "iran"
    if s in CRYPTO_SYMBOLS: return "crypto"
    if s in GOLDAPI_SYMBOLS: return "goldapi"
    return "forex"

def fmt_price(price, unit):
    if price is None: return "---"
    if unit == "IRR": return f"{price:,}"
    if price < 1000: return f"{price:.5f}"
    return f"{price:,.2f}"

async def get_goldapi_price(metal="XAU", currency="USD"):
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(f"https://www.goldapi.io/api/{metal}/{currency}",
                headers={"x-access-token": GOLD_API_KEY})
            data = r.json()
            price = data.get("price")
            change = data.get("chp", 0)
            return price, change, "рҹҹў" if change >= 0 else "рҹ”ҙ"
    except Exception as e:
        print(f"GoldAPI error: {e}")
        return None, 0, "вҡӘ"

async def get_yahoo_price(symbol):
    try:
        async with httpx.AsyncClient(timeout=10,
            headers={"User-Agent": "Mozilla/5.0"}) as client:
            r = await client.get(f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1m&range=1d")
            meta = r.json()["chart"]["result"][0]["meta"]
            price = meta.get("regularMarketPrice", 0)
            prev = meta.get("chartPreviousClose", meta.get("previousClose", price))
            change = ((price - prev) / prev * 100) if prev else 0
            return price, change, "рҹҹў" if change >= 0 else "рҹ”ҙ"
    except:
        return None, 0, "вҡӘ"

async def get_price_by_symbol(symbol, market):
    symbol = symbol.upper()
    try:
        async with httpx.AsyncClient(timeout=10,
            headers={"User-Agent": "Mozilla/5.0"}) as client:
            if market == "iran":
                key = IRAN_SYMBOLS.get(symbol)
                if not key: return None, "ЩҶШ§Щ…ШҙШ®Шө"
                r = await client.get("https://call4.tgju.org/ajax.json")
                val = r.json().get("current", {}).get(key, {}).get("p", "").replace(",", "")
                return int(float(val)), "IRR"
            elif market == "crypto":
                cg_id = CRYPTO_SYMBOLS.get(symbol)
                if not cg_id: return None, "ЩҶШ§Щ…ШҙШ®Шө"
                r = await client.get("https://api.coingecko.com/api/v3/simple/price",
                    params={"ids": cg_id, "vs_currencies": "usd"})
                return r.json().get(cg_id, {}).get("usd"), "USD"
            elif market == "goldapi":
                metal, currency = GOLDAPI_SYMBOLS.get(symbol, ("XAU", "USD"))
                r = await client.get(f"https://www.goldapi.io/api/{metal}/{currency}",
                    headers={"x-access-token": GOLD_API_KEY})
                return r.json().get("price"), "USD"
            else:
                yahoo_sym = FOREX_YAHOO.get(symbol, symbol + "%3DX")
                r = await client.get(f"https://query1.finance.yahoo.com/v8/finance/chart/{yahoo_sym}?interval=1m&range=1d")
                return r.json()["chart"]["result"][0]["meta"].get("regularMarketPrice"), "USD"
    except:
        return None, "ЩҶШ§Щ…ШҙШ®Шө"

async def validate_symbol(symbol):
    market = detect_market(symbol.upper())
    price, unit = await get_price_by_symbol(symbol, market)
    if price and price > 0:
        return True, market, price, unit
    return False, None, None, None

async def get_iran_prices():
    try:
        async with httpx.AsyncClient(timeout=10,
            headers={"User-Agent": "Mozilla/5.0"}) as client:
            current = (await client.get("https://call4.tgju.org/ajax.json")).json().get("current", {})
            def gp(key):
                try: return int(float(current[key]["p"].replace(",", "")))
                except: return None
            return {"dollar": gp("price_dollar_rl"), "gold18": gp("geram18"),
                    "gold_coin": gp("sekee"), "half_coin": gp("nim"),
                    "quarter": gp("rob"), "gram_gold": gp("mesghal"), "silver": gp("silver_999")}
    except Exception as e:
        print(f"Iran prices error: {e}")
        return None

async def get_crypto_prices():
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            data = (await client.get("https://api.coingecko.com/api/v3/simple/price",
                params={"ids": "bitcoin,ethereum,tether,binancecoin,tron",
                        "vs_currencies": "usd", "include_24hr_change": "true"})).json()
            def fmt(cid):
                c = data.get(cid, {})
                p, ch = c.get("usd", 0), c.get("usd_24h_change", 0)
                return p, ch, "рҹҹў" if ch >= 0 else "рҹ”ҙ"
            return {"bitcoin": fmt("bitcoin"), "ethereum": fmt("ethereum"),
                    "tether": fmt("tether"), "bnb": fmt("binancecoin"), "tron": fmt("tron")}
    except Exception as e:
        print(f"Crypto error: {e}")
        return None

async def get_forex_prices():
    try:
        gp, gc, ga = await get_goldapi_price("XAU", "USD")
        sp, sc, sa = await get_goldapi_price("XAG", "USD")
        return {"gold": (gp, gc, ga), "silver": (sp, sc, sa),
                "oil": await get_yahoo_price("BZ%3DF"),
                "eur": await get_yahoo_price("EURUSD%3DX"),
                "gbp": await get_yahoo_price("GBPUSD%3DX")}
    except Exception as e:
        print(f"Forex error: {e}")
        return None

async def check_alarms(app):
    for alarm_id, user_id, symbol, market, target, direction in get_all_active_alarms():
        try:
            price, unit = await get_price_by_symbol(symbol, market)
            if price is None: continue
            if (direction == "above" and price >= target) or (direction == "below" and price <= target):
                arrow = "рҹ“Ҳ" if direction == "above" else "рҹ“ү"
                await app.bot.send_message(user_id,
                    f"рҹ”” ШўЩ„Ш§ШұЩ… ЩҒШ№Ш§Щ„ ШҙШҜ!\n\n{arrow} {symbol}\n"
                    f"ЩӮЫҢЩ…ШӘ ЩҮШҜЩҒ: {fmt_price(target, unit)}\n"
                    f"ЩӮЫҢЩ…ШӘ ЩҒШ№Щ„ЫҢ: {fmt_price(price, unit)} {unit}\n"
                    f"рҹ•җ {datetime.now().strftime('%H:%M:%S')}")
                deactivate_alarm(alarm_id)
        except Exception as e:
            print(f"Alarm error {alarm_id}: {e}")

async def alarm_loop(app):
    while True:
        await check_alarms(app)
        await asyncio.sleep(60)

def financial_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹ’ө ШЁШ§ШІШ§Шұ Ш§ЫҢШұШ§ЩҶ", callback_data="fin_currency")],
        [InlineKeyboardButton("вӮҝ ШЁШ§ШІШ§Шұ Ъ©ШұЫҢЩҫШӘЩҲ", callback_data="fin_crypto")],
        [InlineKeyboardButton("рҹ“Ҳ ШЁШ§ШІШ§Шұ ЩҒШ§ШұЪ©Ші", callback_data="fin_forex")],
        [InlineKeyboardButton("рҹ‘Ғ ЩҲШ§ЪҶ Щ„ЫҢШіШӘ ШҙШ®ШөЫҢ", callback_data="fin_watchlist")],
        [InlineKeyboardButton("рҹ”” ШўЩ„Ш§ШұЩ… ЩӮЫҢЩ…ШӘ", callback_data="fin_alarm")],
        [InlineKeyboardButton("рҹ“җ Щ…ЫҢШІШ§ЩҶ Ш®ШұЫҢШҜ", callback_data="ps_size")],  # вҶҗ ШӯШ°ЩҒ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
        [InlineKeyboardButton("рҹ“Ҡ ШЁЪ©вҖҢШӘШіШӘ", callback_data="fin_backtest")],  # вҶҗ ШӯШ°ЩҒ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_main")],
    ])

def watchlist_menu(user_id):
    items = get_watchlist(user_id)
    keyboard = []
    
    # ШҜЪ©Щ…ЩҮвҖҢЩҮШ§ЫҢ ШӯШ°ЩҒ - 3ШӘШ§ЫҢЫҢ ШҜШұ ЩҮШұ ШіШ·Шұ
    row = []
    for wid, symbol, market in items:
        row.append(InlineKeyboardButton(f"вқҢ {symbol}", callback_data=f"wl_del_{wid}"))
        if len(row) == 3:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    
    keyboard.append([InlineKeyboardButton("вһ• Ш§ЩҒШІЩҲШҜЩҶ Ш§ШұШІ", callback_data="wl_add")])
    keyboard.append([InlineKeyboardButton("рҹ“Ҡ Щ…ШҙШ§ЩҮШҜЩҮ ЩӮЫҢЩ…ШӘвҖҢЩҮШ§", callback_data="wl_view")])
    keyboard.append([InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_financial")])
    return InlineKeyboardMarkup(keyboard)

def alarm_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("вһ• ШўЩ„Ш§ШұЩ… Ш¬ШҜЫҢШҜ", callback_data="alarm_new")],
        [InlineKeyboardButton("рҹ“Ӣ ШўЩ„Ш§ШұЩ…вҖҢЩҮШ§ЫҢ ЩҒШ№Ш§Щ„", callback_data="alarm_list")],
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_financial")],
    ])

async def handle_financial(query, user_id):
    data = query.data

    if data == "menu_financial":
        await query.edit_message_text("рҹ’° ШЁШ§ШІШ§ШұЩҮШ§ЫҢ Щ…Ш§Щ„ЫҢ:", reply_markup=financial_menu())

    elif data == "back_financial":
        clear_state(user_id)
        await query.edit_message_text("рҹ’° ШЁШ§ШІШ§ШұЩҮШ§ЫҢ Щ…Ш§Щ„ЫҢ:", reply_markup=financial_menu())

    elif data == "fin_currency":
        await query.edit_message_text("вҸі ШҜШұ ШӯШ§Щ„ ШҜШұЫҢШ§ЩҒШӘ ЩӮЫҢЩ…ШӘвҖҢЩҮШ§...")
        prices = await get_iran_prices()
        if prices:
            def f(v): return f"{v:,}" if v else "---"
            text = ("рҹ’ө Iran Market\n`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`\n"
                f"`{'USD':<10} {f(prices['dollar']):>15} IRR`\n"
                f"`{'Gold 18':<10} {f(prices['gold18']):>15} IRR`\n"
                f"`{'Coin':<10} {f(prices['gold_coin']):>15} IRR`\n"
                f"`{'Half Coin':<10} {f(prices['half_coin']):>15} IRR`\n"
                f"`{'Qtr Coin':<10} {f(prices['quarter']):>15} IRR`\n"
                f"`{'Mithgal':<10} {f(prices['gram_gold']):>15} IRR`\n"
                f"`{'Silver':<10} {f(prices['silver']):>15} IRR`\n"
                f"`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`\nрҹ•җ {datetime.now().strftime('%H:%M:%S')}")
        else:
            text = "вқҢ Ш®Ш·Ш§ ШҜШұ ШҜШұЫҢШ§ЩҒШӘ ЩӮЫҢЩ…ШӘвҖҢЩҮШ§."
        await query.edit_message_text(text, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_financial")]]))

    elif data == "fin_crypto":
        await query.edit_message_text("вҸі ШҜШұ ШӯШ§Щ„ ШҜШұЫҢШ§ЩҒШӘ ЩӮЫҢЩ…ШӘвҖҢЩҮШ§...")
        prices = await get_crypto_prices()
        if prices:
            def row(n, d):
                p, ch, ar = d
                return f"`{n:<5} ${p:>10,.2f}  {ar} {ch:>+6.2f}%`"
            text = ("вӮҝ Crypto Market\n`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`\n"
                + row("BTC", prices["bitcoin"]) + "\n"
                + row("ETH", prices["ethereum"]) + "\n"
                + row("USDT", prices["tether"]) + "\n"
                + row("BNB", prices["bnb"]) + "\n"
                + row("TRX", prices["tron"]) + "\n"
                f"`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`\nрҹ•җ {datetime.now().strftime('%H:%M:%S')}")
        else:
            text = "вқҢ Ш®Ш·Ш§ ШҜШұ ШҜШұЫҢШ§ЩҒШӘ ЩӮЫҢЩ…ШӘвҖҢЩҮШ§."
        await query.edit_message_text(text, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_financial")]]))

    elif data == "fin_forex":
        await query.edit_message_text("вҸі ШҜШұ ШӯШ§Щ„ ШҜШұЫҢШ§ЩҒШӘ ЩӮЫҢЩ…ШӘвҖҢЩҮШ§...")
        prices = await get_forex_prices()
        if prices:
            def row(n, d):
                p, ch, ar = d
                if p is None: return f"`{n:<8} {'---':>12}`"
                return f"`{n:<8} {fmt_price(p, 'USD'):>12}  {ar} {ch:>+6.2f}%`"
            text = ("рҹ“Ҳ Forex Market\n`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`\n"
                + row("GOLD", prices["gold"]) + "\n"
                + row("SILVER", prices["silver"]) + "\n"
                + row("OIL", prices["oil"]) + "\n"
                + row("EUR/USD", prices["eur"]) + "\n"
                + row("GBP/USD", prices["gbp"]) + "\n"
                f"`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`\nрҹ•җ {datetime.now().strftime('%H:%M:%S')}")
        else:
            text = "вқҢ Ш®Ш·Ш§ ШҜШұ ШҜШұЫҢШ§ЩҒШӘ ЩӮЫҢЩ…ШӘвҖҢЩҮШ§."
        await query.edit_message_text(text, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_financial")]]))

    elif data == "fin_watchlist":
        await query.edit_message_text("рҹ‘Ғ ЩҲШ§ЪҶ Щ„ЫҢШіШӘ ШҙШ®ШөЫҢ:", reply_markup=watchlist_menu(user_id))

    elif data == "wl_add":
        set_state(user_id, "wl_add")
        await query.edit_message_text(
            "вһ• ЩҶЩ…Ш§ШҜ Ш§ШұШІ ШұШ§ ШӘШ§ЫҢЩҫ Ъ©ЩҶЫҢШҜ:\n\nЩ…Ш«Ш§Щ„:\n`BTC` `GOLD` `USDCAD` `EURUSD`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="fin_watchlist")]]))

    elif data == "wl_view":
        items = get_watchlist(user_id)
        if not items:
            await query.answer("ЩҲШ§ЪҶ Щ„ЫҢШіШӘ Ш®Ш§Щ„ЫҢ Ш§ШіШӘ!", show_alert=True)
            return
        await query.edit_message_text("вҸі ШҜШұ ШӯШ§Щ„ ШҜШұЫҢШ§ЩҒШӘ ЩӮЫҢЩ…ШӘвҖҢЩҮШ§...")
        lines = []
        for _, symbol, market in items:
            price, unit = await get_price_by_symbol(symbol, market)
            lines.append(f"`{symbol:<10} {fmt_price(price, unit):>15} {unit}`")
        text = ("рҹ‘Ғ ЩҲШ§ЪҶ Щ„ЫҢШіШӘ Щ…ЩҶ\n`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`\n"
            + "\n".join(lines) + f"\n`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`\nрҹ•җ {datetime.now().strftime('%H:%M:%S')}")
        await query.edit_message_text(text, parse_mode="Markdown", reply_markup=watchlist_menu(user_id))

    elif data.startswith("wl_del_"):
        remove_watchlist(int(data.replace("wl_del_", "")))
        await query.edit_message_text("вң… ШӯШ°ЩҒ ШҙШҜ.", reply_markup=watchlist_menu(user_id))

    elif data == "fin_alarm":
        await query.edit_message_text("рҹ”” ШўЩ„Ш§ШұЩ… ЩӮЫҢЩ…ШӘ:", reply_markup=alarm_menu())

    elif data == "alarm_new":
        set_state(user_id, "alarm_symbol")
        await query.edit_message_text(
            "рҹ”” ЩҶЩ…Ш§ШҜ Ш§ШұШІ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:\n\nЩ…Ш«Ш§Щ„:\n`BTC` `GOLD` `USDCAD`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="fin_alarm")]]))

    elif data == "alarm_list":
        alarms = get_alarms(user_id)
        if not alarms:
            await query.answer("ЩҮЫҢЪҶ ШўЩ„Ш§ШұЩ… ЩҒШ№Ш§Щ„ЫҢ ЩҶШҜШ§ШұЫҢШҜ!", show_alert=True)
            return
        text = "рҹ“Ӣ ШўЩ„Ш§ШұЩ…вҖҢЩҮШ§ЫҢ ЩҒШ№Ш§Щ„:\n\n"
        keyboard = []
        for aid, symbol, market, target, direction in alarms:
            unit = "IRR" if market == "iran" else "USD"
            arrow = "рҹ“Ҳ" if direction == "above" else "рҹ“ү"
            text += f"{arrow} {symbol} вҶҗ {fmt_price(target, unit)} {unit}\n"
            keyboard.append([InlineKeyboardButton(f"рҹ—‘ ШӯШ°ЩҒ {symbol}", callback_data=f"alarm_del_{aid}")])
        keyboard.append([InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="fin_alarm")])
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))

    elif data.startswith("alarm_del_"):
        remove_alarm(int(data.replace("alarm_del_", "")))
        await query.edit_message_text("вң… ШўЩ„Ш§ШұЩ… ШӯШ°ЩҒ ШҙШҜ.", reply_markup=alarm_menu())

    elif data in ["alarm_dir_above", "alarm_dir_below"]:
        state, sdata = get_state(user_id)
        if state == "alarm_direction" and sdata:
            parts = sdata.split("|")
            symbol, market, target, unit = parts[0], parts[1], float(parts[2]), parts[3]
            direction = "above" if data == "alarm_dir_above" else "below"
            add_alarm(user_id, symbol, market, target, direction)
            clear_state(user_id)
            arrow = "рҹ“Ҳ" if direction == "above" else "рҹ“ү"
            await query.edit_message_text(
                f"вң… ШўЩ„Ш§ШұЩ… Ш«ШЁШӘ ШҙШҜ!\n\n{arrow} {symbol}\nЩӮЫҢЩ…ШӘ ЩҮШҜЩҒ: {fmt_price(target, unit)} {unit}",
                reply_markup=alarm_menu())

async def handle_financial_message(user_id, text, update):
    state, data = get_state(user_id)

    if state == "wl_add":
        symbol = text.upper()
        await update.message.reply_text("вҸі ШҜШұ ШӯШ§Щ„ ШЁШұШұШіЫҢ...")
        valid, market, price, unit = await validate_symbol(symbol)
        if valid:
            if add_watchlist(user_id, symbol, market):
                await update.message.reply_text(
                    f"вң… {symbol} Ш§Ш¶Ш§ЩҒЩҮ ШҙШҜ!\nрҹ’° ЩӮЫҢЩ…ШӘ: {fmt_price(price, unit)} {unit}",
                    reply_markup=watchlist_menu(user_id))
            else:
                await update.message.reply_text(f"вҡ пёҸ {symbol} ЩӮШЁЩ„Ш§ЩӢ ЩҮШіШӘ.", reply_markup=watchlist_menu(user_id))
        else:
            await update.message.reply_text(f"вқҢ '{symbol}' ЩҫЫҢШҜШ§ ЩҶШҙШҜ.", reply_markup=watchlist_menu(user_id))
        clear_state(user_id)
        return True

    elif state == "alarm_symbol":
        symbol = text.upper()
        await update.message.reply_text("вҸі ШҜШұ ШӯШ§Щ„ ШЁШұШұШіЫҢ...")
        valid, market, price, unit = await validate_symbol(symbol)
        if valid:
            set_state(user_id, "alarm_price", f"{symbol}|{market}|{price}|{unit}")
            await update.message.reply_text(
                f"вң… {symbol} ЩҫЫҢШҜШ§ ШҙШҜ!\nрҹ’° ЩӮЫҢЩ…ШӘ: {fmt_price(price, unit)} {unit}\n\nЩӮЫҢЩ…ШӘ ЩҮШҜЩҒ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        else:
            await update.message.reply_text(f"вқҢ '{symbol}' ЩҫЫҢШҜШ§ ЩҶШҙШҜ. ШҜЩҲШЁШ§ШұЩҮ Ш§Щ…ШӘШӯШ§ЩҶ Ъ©ЩҶЫҢШҜ:")
        return True

    elif state == "alarm_price":
        try:
            target = float(text.replace(",", ""))
            parts = data.split("|")
            symbol, market, unit = parts[0], parts[1], parts[3]
            set_state(user_id, "alarm_direction", f"{symbol}|{market}|{target}|{unit}")
            await update.message.reply_text("Ш¬ЩҮШӘ ШўЩ„Ш§ШұЩ…:",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton(f"рҹ“Ҳ ШЁШ§Щ„Ш§ШӘШұ Ш§ШІ {fmt_price(target, unit)}", callback_data="alarm_dir_above")],
                    [InlineKeyboardButton(f"рҹ“ү ЩҫШ§ЫҢЫҢЩҶвҖҢШӘШұ Ш§ШІ {fmt_price(target, unit)}", callback_data="alarm_dir_below")],
                ]))
        except:
            await update.message.reply_text("вқҢ Ш№ШҜШҜ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        return True

    return False

# в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ ЩҫШ§ЫҢШ§ЩҶ MODULE: FINANCIAL в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ


# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                  MODULE: ACCOUNTING                          в•‘
# в•‘  ШЁШұШ§ЫҢ ЩҒШ№Ш§Щ„ Ъ©ШұШҜЩҶ: Ъ©ШҜ ШұЩҲ Ш§ЫҢЩҶШ¬Ш§ Ш§Ш¶Ш§ЩҒЩҮ Ъ©ЩҶ                      в•‘
# в•‘  ШЁШұШ§ЫҢ ШӯШ°ЩҒ: Ш§ЫҢЩҶ ШЁЩ„ЩҲЪ© ШұЩҲ ЩҫШ§Ъ© Ъ©ЩҶ                               в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ

def accounting_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("вһ• Ш«ШЁШӘ ШҜШұШўЩ…ШҜ", callback_data="acc_income")],
        [InlineKeyboardButton("вһ– Ш«ШЁШӘ ЩҮШІЫҢЩҶЩҮ", callback_data="acc_expense")],
        [InlineKeyboardButton("рҹ“Ӣ Щ„ЫҢШіШӘ ШӘШұШ§Ъ©ЩҶШҙвҖҢЩҮШ§", callback_data="acc_list")],
        [InlineKeyboardButton("рҹ“Ҳ ЪҜШІШ§ШұШҙ Щ…Ш§ЩҮШ§ЩҶЩҮ", callback_data="acc_report")],
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_main")],
    ])

async def handle_accounting(query, user_id):
    data = query.data
    if data == "menu_accounting":
        await query.edit_message_text("рҹ“Ҡ ШӯШіШ§ШЁШҜШ§ШұЫҢ ШҙШ®ШөЫҢ:", reply_markup=accounting_menu())
    else:
        await query.answer("рҹ”§ ШЁЩҮ ШІЩҲШҜЫҢ Ш§Ш¶Ш§ЩҒЩҮ Щ…ЫҢвҖҢШҙЩҲШҜ!", show_alert=True)

# в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ ЩҫШ§ЫҢШ§ЩҶ MODULE: ACCOUNTING в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ


# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                  MODULE: REMINDER                            в•‘
# в•‘  ШЁШұШ§ЫҢ ШӯШ°ЩҒ: Ш§ЫҢЩҶ ШЁЩ„ЩҲЪ© ШұЩҲ ЩҫШ§Ъ© Ъ©ЩҶ                               в•‘
# в•‘  + Ш®Ш· rem_/menu_reminder ШұЩҲ Ш§ШІ ROUTER Ш§ШөЩ„ЫҢ ЩҫШ§Ъ© Ъ©ЩҶ            в•‘
# в•‘  + state ЩҮШ§ЫҢ rem_ ШұЩҲ Ш§ШІ message_handler ЩҫШ§Ъ© Ъ©ЩҶ              в•‘
# в•‘  + Ш®Ш· reminder_loop ШұЩҲ Ш§ШІ post_init ЩҫШ§Ъ© Ъ©ЩҶ                  в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ

def get_iran_now():
    """ШӘШ§ШұЫҢШ® ЩҲ ШіШ§Ш№ШӘ Ъ©Ш§Щ…Щ„ ЩҒШ№Щ„ЫҢ ШЁЩҮ ЩҲЩӮШӘ Ш§ЫҢШұШ§ЩҶ (datetime object)"""
    from datetime import timezone
    return datetime.now(timezone.utc) + timedelta(hours=3, minutes=30)

# в”Җв”Җв”Җ ШҜЫҢШӘШ§ШЁЫҢШі в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
def init_reminder_db():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS tasks (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id      INTEGER,
        title        TEXT,
        due_date     TEXT,
        due_time     TEXT,
        repeat_type  TEXT DEFAULT 'once',
        alert_before INTEGER DEFAULT 0,
        alert_sent   INTEGER DEFAULT 0,
        due_notified INTEGER DEFAULT 0,
        done         INTEGER DEFAULT 0,
        created_at   TEXT
    )''')
    try:
        c.execute("ALTER TABLE tasks ADD COLUMN due_notified INTEGER DEFAULT 0")
    except:
        pass
    conn.commit()
    conn.close()

def add_task(user_id, title, due_date, due_time, repeat_type, alert_before_minutes):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''INSERT INTO tasks (user_id, title, due_date, due_time, repeat_type, alert_before, created_at)
                 VALUES (?, ?, ?, ?, ?, ?, ?)''',
              (user_id, title, due_date, due_time, repeat_type, alert_before_minutes,
               get_iran_now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def get_user_tasks(user_id, only_pending=True):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    if only_pending:
        c.execute('''SELECT id, title, due_date, due_time, repeat_type, alert_before, done
                     FROM tasks WHERE user_id=? AND done=0 ORDER BY due_date, due_time''', (user_id,))
    else:
        c.execute('''SELECT id, title, due_date, due_time, repeat_type, alert_before, done
                     FROM tasks WHERE user_id=? AND done=1 ORDER BY due_date DESC, due_time DESC''', (user_id,))
    rows = c.fetchall()
    conn.close()
    return rows

def get_tasks_for_date(user_id, date_str):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''SELECT id, title, due_time FROM tasks
                 WHERE user_id=? AND due_date=? AND done=0 ORDER BY due_time''', (user_id, date_str))
    rows = c.fetchall()
    conn.close()
    return rows

def mark_task_done(task_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE tasks SET done=1 WHERE id=?", (task_id,))
    conn.commit()
    conn.close()

def delete_task(task_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("DELETE FROM tasks WHERE id=?", (task_id,))
    conn.commit()
    conn.close()

def mark_alert_sent(task_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE tasks SET alert_sent=1 WHERE id=?", (task_id,))
    conn.commit()
    conn.close()

def mark_due_notified(task_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE tasks SET due_notified=1 WHERE id=?", (task_id,))
    conn.commit()
    conn.close()

def advance_recurring_task(task_id, repeat_type, due_date):
    """ШӘШіЪ© ШӘЪ©ШұШ§ШұШҙЩҲЩҶШҜЩҮ ШұШ§ ШЁЩҮ ШҜЩҲШұ ШЁШ№ШҜЫҢ Щ…ЩҶШӘЩӮЩ„ Щ…ЫҢвҖҢЪ©ЩҶШҜ (ШӘШ§ШұЫҢШ® Ш¬ШҜЫҢШҜ + ШұЫҢШіШӘ ЩҲШ¶Ш№ЫҢШӘ ЩҮШҙШҜШ§Шұ)"""
    current = datetime.strptime(due_date, "%Y-%m-%d")
    if repeat_type == "daily":
        next_date = current + timedelta(days=1)
    elif repeat_type == "weekly":
        next_date = current + timedelta(days=7)
    else:
        return
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE tasks SET due_date=?, alert_sent=0, due_notified=0, done=0 WHERE id=?",
              (next_date.strftime("%Y-%m-%d"), task_id))
    conn.commit()
    conn.close()

def get_all_pending_tasks_raw():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''SELECT id, user_id, title, due_date, due_time, repeat_type, alert_before, alert_sent, due_notified
                 FROM tasks WHERE done=0''')
    rows = c.fetchall()
    conn.close()
    return rows

def get_morning_summary_users():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT DISTINCT user_id FROM tasks WHERE done=0")
    rows = [r[0] for r in c.fetchall()]
    conn.close()
    return rows

# в”Җв”Җв”Җ Щ…ЩҶЩҲЩҮШ§ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
def reminder_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("вһ• ШӘШіЪ© Ш¬ШҜЫҢШҜ", callback_data="rem_new")],
        [InlineKeyboardButton("рҹ“Ӣ Щ„ЫҢШіШӘ ШӘШіЪ©вҖҢЩҮШ§", callback_data="rem_list")],
        [InlineKeyboardButton("вң… Ш§ЩҶШ¬Ш§Щ… ШҙШҜЩҮвҖҢЩҮШ§", callback_data="rem_done_list")],
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_main")],
    ])

def repeat_type_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹ”Ӯ ЫҢЪ©вҖҢШЁШ§Шұ", callback_data="rem_rep_once")],
        [InlineKeyboardButton("рҹ”Ғ ЩҮШұ ШұЩҲШІ", callback_data="rem_rep_daily")],
        [InlineKeyboardButton("рҹ“… ЩҮШұ ЩҮЩҒШӘЩҮ", callback_data="rem_rep_weekly")],
    ])

def alert_yes_no_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("вң… ШЁЩ„ЩҮ", callback_data="rem_alert_yes")],
        [InlineKeyboardButton("вқҢ Ш®ЫҢШұ", callback_data="rem_alert_no")],
    ])

def alert_unit_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("вҸұ ШҜЩӮЫҢЩӮЩҮ", callback_data="rem_unit_minute")],
        [InlineKeyboardButton("рҹ•җ ШіШ§Ш№ШӘ", callback_data="rem_unit_hour")],
    ])

REPEAT_LABELS = {"once": "ЫҢЪ©вҖҢШЁШ§Шұ", "daily": "ЩҮШұ ШұЩҲШІ", "weekly": "ЩҮШұ ЩҮЩҒШӘЩҮ"}

# в”Җв”Җв”Җ ЩҮЩҶШҜЩ„Шұ ШҜЪ©Щ…ЩҮвҖҢЩҮШ§ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
async def handle_reminder(query, user_id):
    data = query.data

    if data == "menu_reminder":
        await query.edit_message_text("вҸ° ЫҢШ§ШҜШўЩҲШұ ЩҲ ШӘШіЪ©:", reply_markup=reminder_menu())

    elif data == "rem_new":
        set_state(user_id, "rem_title")
        await query.edit_message_text(
            "вһ• ШӘШіЪ© Ш¬ШҜЫҢШҜ\n\nШ№ЩҶЩҲШ§ЩҶ ШӘШіЪ© ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="menu_reminder")]])
        )

    elif data == "rem_list":
        await handle_reminder_refresh_list(query, user_id)

    elif data == "rem_done_list":
        tasks = get_user_tasks(user_id, only_pending=False)
        if not tasks:
            await query.edit_message_text("рҹ“ӯ ЩҮЫҢЪҶ ШӘШіЪ© Ш§ЩҶШ¬Ш§Щ…вҖҢШҙШҜЩҮвҖҢШ§ЫҢ ЩҶШҜШ§ШұЫҢШҜ.", reply_markup=reminder_menu())
            return
        text = "вң… ШӘШіЪ©вҖҢЩҮШ§ЫҢ Ш§ЩҶШ¬Ш§Щ…вҖҢШҙШҜЩҮ (ЫұЫ° Щ…ЩҲШұШҜ ШўШ®Шұ):\n\n"
        for tid, title, due_date, due_time, repeat_type, alert_before, done in tasks[:10]:
            text += f"вң“ {title} вҖ” {due_date}\n"
        await query.edit_message_text(
            text, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="menu_reminder")]])
        )

    elif data.startswith("rem_done_"):
        tid = int(data.replace("rem_done_", ""))
        mark_task_done(tid)
        await query.answer("вң… ШӘШіЪ© Ш§ЩҶШ¬Ш§Щ…вҖҢШҙШҜЩҮ Ш«ШЁШӘ ШҙШҜ.", show_alert=True)
        await handle_reminder_refresh_list(query, user_id)

    elif data.startswith("rem_del_"):
        tid = int(data.replace("rem_del_", ""))
        delete_task(tid)
        await query.answer("рҹ—‘ ШӘШіЪ© ШӯШ°ЩҒ ШҙШҜ.", show_alert=True)
        await handle_reminder_refresh_list(query, user_id)

    elif data.startswith("rem_rep_"):
        repeat_type = data.replace("rem_rep_", "")
        state, sdata = get_state(user_id)
        if state == "rem_repeat":
            set_state(user_id, "rem_alert_choice", f"{sdata}|{repeat_type}")
            await query.edit_message_text(
                "рҹ”” Щ…ЫҢвҖҢШ®ЩҲШ§ЩҮЫҢШҜ ЩӮШЁЩ„ Ш§ШІ Щ…ЩҲШ№ШҜШҢ ЫҢШ§ШҜШўЩҲШұЫҢ ЩҮШҙШҜШ§Шұ ШЁШұШ§ЫҢШӘШ§ЩҶ Ш§ШұШіШ§Щ„ ШҙЩҲШҜШҹ",
                reply_markup=alert_yes_no_menu()
            )

    elif data in ["rem_alert_yes", "rem_alert_no"]:
        state, sdata = get_state(user_id)
        if state == "rem_alert_choice":
            if data == "rem_alert_no":
                await finalize_task_creation(query, user_id, sdata, alert_before_minutes=0)
            else:
                set_state(user_id, "rem_alert_unit", sdata)
                await query.edit_message_text(
                    "вҸі ЩҲШ§ШӯШҜ ШІЩ…Ш§ЩҶЫҢ ЩҮШҙШҜШ§Шұ ШұШ§ Ш§ЩҶШӘШ®Ш§ШЁ Ъ©ЩҶЫҢШҜ:",
                    reply_markup=alert_unit_menu()
                )

    elif data in ["rem_unit_minute", "rem_unit_hour"]:
        state, sdata = get_state(user_id)
        if state == "rem_alert_unit":
            unit = "minute" if data == "rem_unit_minute" else "hour"
            set_state(user_id, "rem_alert_value", f"{sdata}|{unit}")
            unit_fa = "ШҜЩӮЫҢЩӮЩҮ" if unit == "minute" else "ШіШ§Ш№ШӘ"
            await query.edit_message_text(f"ЪҶЩҶШҜ {unit_fa} ЩӮШЁЩ„ Ш§ШІ Щ…ЩҲШ№ШҜ ЩҮШҙШҜШ§Шұ ШЁШҜЩҮЩ…Шҹ (Ш№ШҜШҜ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ)")

async def handle_reminder_refresh_list(query, user_id):
    tasks = get_user_tasks(user_id, only_pending=True)
    if not tasks:
        await query.edit_message_text("рҹ“ӯ ЩҮЫҢЪҶ ШӘШіЪ© ЩҒШ№Ш§Щ„ЫҢ ЩҶШҜШ§ШұЫҢШҜ.", reply_markup=reminder_menu())
        return
    text = "рҹ“Ӣ ШӘШіЪ©вҖҢЩҮШ§ЫҢ ЩҒШ№Ш§Щ„:\n\n"
    keyboard = []
    for tid, title, due_date, due_time, repeat_type, alert_before, done in tasks:
        rep_label = REPEAT_LABELS.get(repeat_type, "ЫҢЪ©вҖҢШЁШ§Шұ")
        text += f"рҹ”ё {title}\nрҹ“… {due_date} вҸ° {due_time} ({rep_label})\n\n"
        keyboard.append([
            InlineKeyboardButton(f"вң… {title[:15]}", callback_data=f"rem_done_{tid}"),
            InlineKeyboardButton("рҹ—‘", callback_data=f"rem_del_{tid}")
        ])
    keyboard.append([InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="menu_reminder")])
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))

async def finalize_task_creation(query, user_id, sdata, alert_before_minutes):
    parts = sdata.split("|")
    title, due_date, due_time, repeat_type = parts[0], parts[1], parts[2], parts[3]
    add_task(user_id, title, due_date, due_time, repeat_type, alert_before_minutes)
    clear_state(user_id)
    rep_label = REPEAT_LABELS.get(repeat_type, "ЫҢЪ©вҖҢШЁШ§Шұ")
    alert_line = f"\nрҹ”” ЩҮШҙШҜШ§Шұ: {alert_before_minutes} ШҜЩӮЫҢЩӮЩҮ ЩӮШЁЩ„" if alert_before_minutes > 0 else ""
    await query.edit_message_text(
        f"вң… ШӘШіЪ© Ш«ШЁШӘ ШҙШҜ!\n\n"
        f"рҹ”ё {title}\nрҹ“… {due_date} вҸ° {due_time}\nрҹ”Ғ ШӘЪ©ШұШ§Шұ: {rep_label}{alert_line}",
        reply_markup=reminder_menu()
    )

async def handle_reminder_message(user_id, text, update):
    state, data = get_state(user_id)
    text = text.strip()

    if state == "rem_title":
        set_state(user_id, "rem_date", text)
        await update.message.reply_text("рҹ“… ШӘШ§ШұЫҢШ® Ш§ЩҶШ¬Ш§Щ… ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ (ЩҒШұЩ…ШӘ: YYYY-MM-DD):\nЩ…Ш«Ш§Щ„: `2026-06-28`", parse_mode="Markdown")
        return True

    elif state == "rem_date":
        try:
            datetime.strptime(text, "%Y-%m-%d")
            set_state(user_id, "rem_time", f"{data}|{text}")
            await update.message.reply_text("вҸ° ШіШ§Ш№ШӘ Ш§ЩҶШ¬Ш§Щ… ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ (ЩҒШұЩ…ШӘ 24 ШіШ§Ш№ШӘЩҮ HH:MM):\nЩ…Ш«Ш§Щ„: `14:30`", parse_mode="Markdown")
        except:
            await update.message.reply_text("вқҢ ЩҒШұЩ…ШӘ ШӘШ§ШұЫҢШ® Ш§ШҙШӘШЁШ§ЩҮ Ш§ШіШӘ.\nЩ…Ш«Ш§Щ„: `2026-06-28`", parse_mode="Markdown")
        return True

    elif state == "rem_time":
        try:
            datetime.strptime(text, "%H:%M")
            set_state(user_id, "rem_repeat", f"{data}|{text}")
            await update.message.reply_text("рҹ”Ғ ЩҶЩҲШ№ ШӘЪ©ШұШ§Шұ ШӘШіЪ© ШұШ§ Ш§ЩҶШӘШ®Ш§ШЁ Ъ©ЩҶЫҢШҜ:", reply_markup=repeat_type_menu())
        except:
            await update.message.reply_text("вқҢ ЩҒШұЩ…ШӘ ШіШ§Ш№ШӘ Ш§ШҙШӘШЁШ§ЩҮ Ш§ШіШӘ.\nЩ…Ш«Ш§Щ„: `14:30`", parse_mode="Markdown")
        return True

    elif state == "rem_alert_value":
        try:
            value = int(text)
            if value <= 0:
                await update.message.reply_text("вқҢ Ш№ШҜШҜ ШЁШ§ЫҢШҜ ШЁШІШұЪҜвҖҢШӘШұ Ш§ШІ ШөЩҒШұ ШЁШ§ШҙШҜ:")
                return True
            parts = data.split("|")
            title, due_date, due_time, repeat_type, unit = parts[0], parts[1], parts[2], parts[3], parts[4]
            alert_minutes = value if unit == "minute" else value * 60

            add_task(user_id, title, due_date, due_time, repeat_type, alert_minutes)
            clear_state(user_id)
            rep_label = REPEAT_LABELS.get(repeat_type, "ЫҢЪ©вҖҢШЁШ§Шұ")
            unit_fa = "ШҜЩӮЫҢЩӮЩҮ" if unit == "minute" else "ШіШ§Ш№ШӘ"
            await update.message.reply_text(
                f"вң… ШӘШіЪ© Ш«ШЁШӘ ШҙШҜ!\n\n"
                f"рҹ”ё {title}\nрҹ“… {due_date} вҸ° {due_time}\nрҹ”Ғ ШӘЪ©ШұШ§Шұ: {rep_label}\n"
                f"рҹ”” ЩҮШҙШҜШ§Шұ: {value} {unit_fa} ЩӮШЁЩ„ Ш§ШІ Щ…ЩҲШ№ШҜ",
                reply_markup=reminder_menu()
            )
        except:
            await update.message.reply_text("вқҢ Щ„Ш·ЩҒШ§ЩӢ ЫҢЪ© Ш№ШҜШҜ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        return True

    return False

# в”Җв”Җв”Җ ШӯЩ„ЩӮЩҮ ЩҫШівҖҢШІЩ…ЫҢЩҶЩҮ: ЩҮШҙШҜШ§ШұШҢ ЫҢШ§ШҜШўЩҲШұЫҢ ШІЩ…Ш§ЩҶШҢ Ш®Щ„Ш§ШөЩҮ ШөШЁШӯЪҜШ§ЩҮЫҢ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
async def check_reminders(app):
    now_iran = get_iran_now()
    today_str = now_iran.strftime("%Y-%m-%d")
    now_time_str = now_iran.strftime("%H:%M")
    now_dt = datetime.strptime(f"{today_str} {now_time_str}", "%Y-%m-%d %H:%M")

    tasks = get_all_pending_tasks_raw()
    for tid, user_id, title, due_date, due_time, repeat_type, alert_before, alert_sent, due_notified in tasks:
        try:
            due_dt = datetime.strptime(f"{due_date} {due_time}", "%Y-%m-%d %H:%M")

            # ЩҮШҙШҜШ§Шұ ЩҫЫҢШҙ Ш§ШІ Щ…ЩҲШ№ШҜ (ШЁШ§ШІЩҮвҖҢШ§ЫҢШҢ ЩҶЩҮ ШӘШіШ§ЩҲЫҢ ШҜЩӮЫҢЩӮ - Щ…ЩӮШ§ЩҲЩ… ШҜШұ ШЁШұШ§ШЁШұ ШӘШЈШ®ЫҢШұ ШӯЩ„ЩӮЩҮ)
            if alert_before > 0 and not alert_sent:
                alert_dt = due_dt - timedelta(minutes=alert_before)
                if alert_dt <= now_dt < due_dt:
                    await app.bot.send_message(
                        user_id,
                        f"рҹ”” ЫҢШ§ШҜШўЩҲШұЫҢ ШІЩҲШҜЩҮЩҶЪҜШ§Щ…!\n\nрҹ”ё {title}\nвҸ° Щ…ЩҲШ№ШҜ: {due_time} Ш§Щ…ШұЩҲШІ\n"
                        f"({alert_before} ШҜЩӮЫҢЩӮЩҮ ШҜЫҢЪҜШұ ЩҒШұШ§ Щ…ЫҢвҖҢШұШіШҜ)"
                    )
                    mark_alert_sent(tid)

            # ШұШіЫҢШҜЩҶ ШЁЩҮ Ш®ЩҲШҜ Щ…ЩҲШ№ШҜ ШӘШіЪ© - ШЁШ§ШІЩҮвҖҢШ§ЫҢ (Ш§ШІ Щ„ШӯШёЩҮ Щ…ЩҲШ№ШҜ ШӘШ§ Ыө ШҜЩӮЫҢЩӮЩҮ ШЁШ№ШҜ) ШӘШ§ Ш§ЪҜШұ
            # ШӯЩ„ЩӮЩҮ ШҜЩӮЫҢЩӮШ§ЩӢ ЩҮЩ…Ш§ЩҶ ШҜЩӮЫҢЩӮЩҮ ЪҶЪ© ЩҶЪ©ШұШҜШҢ ЩҫЫҢШ§Щ… Ш§ШІ ЩӮЩ„Щ… ЩҶЫҢЩҒШӘШҜШӣ ШЁШ§ ЩҒЩ„ЪҜ due_notified
            # Ш§ШІ Ш§ШұШіШ§Щ„ ШӘЪ©ШұШ§ШұЫҢ Ш¬Щ„ЩҲЪҜЫҢШұЫҢ Щ…ЫҢвҖҢШҙЩҲШҜ
            if not due_notified and due_dt <= now_dt < due_dt + timedelta(minutes=5):
                await app.bot.send_message(user_id, f"вҸ° ЩҲЩӮШӘШҙЩҮ!\n\nрҹ”ё {title}")
                mark_due_notified(tid)
                if repeat_type in ["daily", "weekly"]:
                    advance_recurring_task(tid, repeat_type, due_date)
                else:
                    mark_task_done(tid)
        except Exception as e:
            print(f"Ш®Ш·Ш§ЫҢ ШЁШұШұШіЫҢ ШӘШіЪ© {tid}: {e}")

async def send_morning_summary(app):
    today_str = get_iran_now().strftime("%Y-%m-%d")
    for user_id in get_morning_summary_users():
        tasks_today = get_tasks_for_date(user_id, today_str)
        if not tasks_today:
            continue
        text = "рҹҢ… Ш®Щ„Ш§ШөЩҮ ШөШЁШӯЪҜШ§ЩҮЫҢ вҖ” Ъ©Ш§ШұЩҮШ§ЫҢ Ш§Щ…ШұЩҲШІ:\n\n"
        for tid, title, due_time in tasks_today:
            text += f"рҹ”ё {due_time} вҖ” {title}\n"
        try:
            await app.bot.send_message(user_id, text)
        except Exception as e:
            print(f"Ш®Ш·Ш§ЫҢ Ш§ШұШіШ§Щ„ Ш®Щ„Ш§ШөЩҮ ШөШЁШӯЪҜШ§ЩҮЫҢ ШЁЩҮ {user_id}: {e}")

async def reminder_loop(app):
    last_summary_date = None
    while True:
        try:
            now_iran = get_iran_now()
            await check_reminders(app)

            # Ш®Щ„Ш§ШөЩҮ ШөШЁШӯЪҜШ§ЩҮЫҢ ШҜЩӮЫҢЩӮШ§ЩӢ ШіШ§Ш№ШӘ Ыё:Ы°Ы° ШӘШ§ Ыё:Ы°Ыұ ШЁЩҮ ЩҲЩӮШӘ Ш§ЫҢШұШ§ЩҶШҢ ЩҒЩӮШ· ЫҢЪ©вҖҢШЁШ§Шұ ШҜШұ ШұЩҲШІ
            if now_iran.strftime("%H:%M") == "08:00":
                today_str = now_iran.strftime("%Y-%m-%d")
                if last_summary_date != today_str:
                    await send_morning_summary(app)
                    last_summary_date = today_str
        except Exception as e:
            print(f"Ш®Ш·Ш§ЫҢ ШӯЩ„ЩӮЩҮ ЫҢШ§ШҜШўЩҲШұ: {e}")
        await asyncio.sleep(60)

# в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ ЩҫШ§ЫҢШ§ЩҶ MODULE: REMINDER в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ



# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                  MODULE: AI CHAT                             в•‘
# в•‘  ЪҶШӘ ШЁШ§ ЪҶЩҶШҜ Щ…ЩҶШЁШ№ ЩҮЩҲШҙ Щ…ШөЩҶЩҲШ№ЫҢ (Groq Ш§ШөЩ„ЫҢШҢ Gemini ШҜШұ ШөЩҲШұШӘ        в•‘
# в•‘  ШҜШ§ШҙШӘЩҶ Ъ©Щ„ЫҢШҜ ШЁЩҮвҖҢШ№ЩҶЩҲШ§ЩҶ ЩҫШҙШӘЫҢШЁШ§ЩҶ) + Щ…ШӯШҜЩҲШҜЫҢШӘ ШұЩҲШІШ§ЩҶЩҮ ЩҲЫҢШұШ§ЫҢШҙ Ш№Ъ©Ші    в•‘
# в•‘  ШЁШұШ§ЫҢ ШӯШ°ЩҒ: Ш§ЫҢЩҶ ШЁЩ„ЩҲЪ© ШұЩҲ ЩҫШ§Ъ© Ъ©ЩҶ                                в•‘
# в•‘  + Ш®Ш· menu_ai ШұЩҲ Ш§ШІ main_menu ЩҫШ§Ъ© Ъ©ЩҶ                        в•‘
# в•‘  + Ш®Ш· ai_ ШұЩҲ Ш§ШІ ROUTER ЩҫШ§Ъ© Ъ©ЩҶ                                в•‘
# в•‘  + state ЩҮШ§ЫҢ ai_ ШұЩҲ Ш§ШІ message_handler ЩҫШ§Ъ© Ъ©ЩҶ               в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ

GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")  # Ш§Ш®ШӘЫҢШ§ШұЫҢ - ЩҲЩӮШӘЫҢ ЪҜШұЩҒШӘЫҢ ЩҮЩ…ЫҢЩҶШ¬Ш§ Ш§Ш¶Ш§ЩҒЩҮ Щ…ЫҢвҖҢШҙЩҲШҜ

GROQ_MODEL = "openai/gpt-oss-120b"
GEMINI_MODEL = "gemini-2.0-flash"

IMAGE_EDIT_DAILY_LIMIT = 4

# в”Җв”Җв”Җ ШҜЫҢШӘШ§ШЁЫҢШі Ш§ЫҢЩҶ Щ…Ш§ЪҳЩҲЩ„ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
def init_ai_db():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS ai_image_usage (
        user_id    INTEGER PRIMARY KEY,
        used_count INTEGER DEFAULT 0,
        reset_date TEXT
    )''')
    conn.commit()
    conn.close()

def get_iran_today_str():
    """ШӘШ§ШұЫҢШ® Ш§Щ…ШұЩҲШІ ШЁЩҮ ЩҲЩӮШӘ Ш§ЫҢШұШ§ЩҶ (ШЁШұШ§ЫҢ ШӘШҙШ®ЫҢШө ШұЫҢШіШӘ ЩҶЫҢЩ…ЩҮвҖҢШҙШЁ Ш§ЫҢШұШ§ЩҶ)"""
    from datetime import timezone
    iran_now = datetime.now(timezone.utc) + timedelta(hours=3, minutes=30)
    return iran_now.strftime("%Y-%m-%d")

def get_image_usage(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT used_count, reset_date FROM ai_image_usage WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    today = get_iran_today_str()
    if not row or row[1] != today:
        return 0, today
    return row[0], row[1]

def increment_image_usage(user_id):
    used, today = get_image_usage(user_id)
    used += 1
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''INSERT INTO ai_image_usage (user_id, used_count, reset_date) VALUES (?, ?, ?)
                 ON CONFLICT(user_id) DO UPDATE SET used_count=excluded.used_count, reset_date=excluded.reset_date''',
              (user_id, used, today))
    conn.commit()
    conn.close()
    return used

# в”Җв”Җв”Җ ЩҒШұШ§Ш®ЩҲШ§ЩҶЫҢ Groq (Щ…ЩҶШЁШ№ Ш§ШөЩ„ЫҢ ЪҶШӘ) в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
async def call_groq_chat(messages):
    if not GROQ_API_KEY:
        return None
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization": f"Bearer {GROQ_API_KEY}"},
                json={"model": GROQ_MODEL, "messages": messages, "temperature": 0.7}
            )
            data = r.json()
            if "choices" in data:
                return data["choices"][0]["message"]["content"]
            print(f"Groq error response: {data}")
            return None
    except Exception as e:
        print(f"Ш®Ш·Ш§ЫҢ Groq: {e}")
        return None

# в”Җв”Җв”Җ ЩҒШұШ§Ш®ЩҲШ§ЩҶЫҢ Gemini (Щ…ЩҶШЁШ№ ЩҫШҙШӘЫҢШЁШ§ЩҶШҢ ЩҒШ№Ш§Щ„ ЩҒЩӮШ· ШҜШұ ШөЩҲШұШӘ ШҜШ§ШҙШӘЩҶ Ъ©Щ„ЫҢШҜ) в”Җ
async def call_gemini_chat(messages):
    if not GEMINI_API_KEY:
        return None
    try:
        # ШӘШЁШҜЫҢЩ„ ЩҒШұЩ…ШӘ ЩҫЫҢШ§Щ…вҖҢЩҮШ§ЫҢ OpenAI-style ШЁЩҮ ЩҒШұЩ…ШӘ Gemini
        contents = []
        for m in messages:
            if m["role"] == "system":
                continue
            role = "user" if m["role"] == "user" else "model"
            contents.append({"role": role, "parts": [{"text": m["content"]}]})

        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent",
                params={"key": GEMINI_API_KEY},
                json={"contents": contents}
            )
            data = r.json()
            candidates = data.get("candidates", [])
            if candidates:
                return candidates[0]["content"]["parts"][0]["text"]
            print(f"Gemini error response: {data}")
            return None
    except Exception as e:
        print(f"Ш®Ш·Ш§ЫҢ Gemini: {e}")
        return None

async def get_ai_response(messages):
    """ШӘЩ„Ш§Шҙ ШЁШ§ GroqШҢ ШҜШұ ШөЩҲШұШӘ Ш®Ш·Ш§ ШӘЩ„Ш§Шҙ ШЁШ§ Gemini (Ш§ЪҜШұ Ъ©Щ„ЫҢШҜШҙ Щ…ЩҲШ¬ЩҲШҜ ШЁШ§ШҙШҜ)"""
    response = await call_groq_chat(messages)
    if response:
        return response
    response = await call_gemini_chat(messages)
    if response:
        return response
    return None

# в”Җв”Җв”Җ Щ…ЩҶЩҲЩҮШ§ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
def ai_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹ’¬ ШҙШұЩҲШ№ ЪҶШӘ", callback_data="ai_start")],
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_main")],
    ])

# в”Җв”Җв”Җ ЩҮЩҶШҜЩ„Шұ ШҜЪ©Щ…ЩҮвҖҢЩҮШ§ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
async def handle_ai(query, user_id):
    data = query.data
    if data == "menu_ai":
        await query.edit_message_text(
            "рҹӨ– ШҜШіШӘЫҢШ§Шұ ЩҮЩҲШҙ Щ…ШөЩҶЩҲШ№ЫҢ\n\nЩҮШұ ШіЩҲШ§Щ„ЫҢ ШҜШ§ШұЫҢ ШЁЩҫШұШіШҢ ШЁШҜЩҲЩҶ Щ…ШӯШҜЩҲШҜЫҢШӘ ШӘШ№ШҜШ§ШҜ.",
            reply_markup=ai_menu()
        )
    elif data == "ai_start":
        set_state(user_id, "ai_chatting")
        await query.edit_message_text(
            "рҹ’¬ ЪҶШӘ ШЁШ§ ШҜШіШӘЫҢШ§Шұ ЩҮЩҲШҙ Щ…ШөЩҶЩҲШ№ЫҢ\n\nШіЩҲШ§Щ„ Ш®ЩҲШҜ ШұШ§ ШЁЩҶЩҲЫҢШіЫҢШҜ:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("рҹ”ҷ ЩҫШ§ЫҢШ§ЩҶ ЪҶШӘ", callback_data="menu_ai")]])
        )

async def handle_ai_message(user_id, text, update):
    state, _ = get_state(user_id)
    if state != "ai_chatting":
        return False

    await update.message.reply_text("вҸі ШҜШұ ШӯШ§Щ„ ЩҒЪ©Шұ Ъ©ШұШҜЩҶ...")
    messages = [
        {"role": "system", "content": "ШӘЩҲ ЫҢЪ© ШҜШіШӘЫҢШ§Шұ ЩҮЩҲШҙЩ…ЩҶШҜ ЩҲ Щ…ЩҒЫҢШҜ ЩҮШіШӘЫҢ Ъ©ЩҮ ШЁЩҮ ШІШЁШ§ЩҶ ЩҒШ§ШұШіЫҢ ЩҫШ§ШіШ® Щ…ЫҢвҖҢШҜЩҮЫҢ."},
        {"role": "user", "content": text}
    ]
    response = await get_ai_response(messages)

    if response:
        await update.message.reply_text(
            response,
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("рҹ”ҷ ЩҫШ§ЫҢШ§ЩҶ ЪҶШӘ", callback_data="menu_ai")]])
        )
    else:
        await update.message.reply_text(
            "вқҢ Щ…ШӘШЈШіЩҒШ§ЩҶЩҮ ШҜШіШӘЫҢШ§Шұ ЩҮЩҲШҙ Щ…ШөЩҶЩҲШ№ЫҢ Щ…ЩҲЩӮШӘШ§ЩӢ ШҜШұ ШҜШіШӘШұШі ЩҶЫҢШіШӘ. Ъ©Щ…ЫҢ ШЁШ№ШҜ ШҜЩҲШЁШ§ШұЩҮ Ш§Щ…ШӘШӯШ§ЩҶ Ъ©ЩҶЫҢШҜ.",
            reply_markup=ai_menu()
        )
    return True

# в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ ЩҫШ§ЫҢШ§ЩҶ MODULE: AI CHAT в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ


# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                  MODULE: KIDS                                в•‘
# в•‘  ЩӮШөЩҮвҖҢЪҜЩҲЫҢЫҢ ЩҮЩҲШҙЩ…ЩҶШҜ Щ…ШӘЩҶШ§ШіШЁ ШЁШ§ ШіЩҶ ЩҲ Ш¬ЩҶШіЫҢШӘ Ъ©ЩҲШҜЪ© + Ш®ШұЩҲШ¬ЫҢ PDF       в•‘
# в•‘  Ш§ШІ ЩҮЩ…Ш§ЩҶ Щ…ЩҶШЁШ№ AI Щ…Ш§ЪҳЩҲЩ„ ЩӮШЁЩ„ЫҢ (Groq/Gemini) Ш§ШіШӘЩҒШ§ШҜЩҮ Щ…ЫҢвҖҢЪ©ЩҶШҜ      в•‘
# в•‘  ШЁШұШ§ЫҢ ШӯШ°ЩҒ: Ш§ЫҢЩҶ ШЁЩ„ЩҲЪ© ШұЩҲ ЩҫШ§Ъ© Ъ©ЩҶ                                в•‘
# в•‘  + Ш®Ш· menu_kids ШұЩҲ Ш§ШІ main_menu ЩҫШ§Ъ© Ъ©ЩҶ                      в•‘
# в•‘  + Ш®Ш· kids_ ШұЩҲ Ш§ШІ ROUTER ЩҫШ§Ъ© Ъ©ЩҶ                              в•‘
# в•‘  + state ЩҮШ§ЫҢ kids_ ШұЩҲ Ш§ШІ message_handler ЩҫШ§Ъ© Ъ©ЩҶ             в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ

def kids_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹ“– ЩӮШөЩҮ Ш¬ШҜЫҢШҜ", callback_data="kids_story")],
        [InlineKeyboardButton("вңҸпёҸ ШӘЪ©Ш§Щ„ЫҢЩҒ ЩҲ Щ…Ш·Ш§Щ„Ш№ЩҮ", callback_data="kids_homework")],
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_main")],
    ])

def kids_gender_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹ‘§ ШҜШ®ШӘШұ", callback_data="kids_gender_girl")],
        [InlineKeyboardButton("рҹ‘Ұ ЩҫШіШұ", callback_data="kids_gender_boy")],
    ])

def build_story_prompt(age, gender, topic):
    gender_fa = "ШҜШ®ШӘШұ" if gender == "girl" else "ЩҫШіШұ"
    return (
        f"ЫҢЪ© ШҜШ§ШіШӘШ§ЩҶ Ъ©ЩҲШҜЪ©Ш§ЩҶЩҮ ШЁЩҮ ШІШЁШ§ЩҶ ЩҒШ§ШұШіЫҢ ШЁШұШ§ЫҢ ЫҢЪ© {gender_fa} {age} ШіШ§Щ„ЩҮ ШЁЩҶЩҲЫҢШі.\n\n"
        f"Щ…ЩҲШ¶ЩҲШ№ ШҜШ§ШіШӘШ§ЩҶ: {topic}\n\n"
        f"ЩӮЩҲШ§ЩҶЫҢЩҶ Щ…ЩҮЩ…:\n"
        f"- ШҜШ§ШіШӘШ§ЩҶ ШЁШ§ЫҢШҜ Ъ©Ш§Щ…Щ„Ш§ЩӢ Щ…ШӘЩҶШ§ШіШЁ ШЁШ§ ШіЩҶ {age} ШіШ§Щ„ ШЁШ§ШҙШҜ (ЩҲШ§ЪҳЪҜШ§ЩҶ ШіШ§ШҜЩҮ ЩҲ ЩӮШ§ШЁЩ„вҖҢЩҒЩҮЩ… ШЁШұШ§ЫҢ Ш§ЫҢЩҶ ШіЩҶ)\n"
        f"- ШҜШ§ШіШӘШ§ЩҶ ШЁШ§ЫҢШҜ ШўЩ…ЩҲШІЩҶШҜЩҮ ШЁШ§ШҙШҜ ЩҲ ЫҢЪ© ЩҶШӘЫҢШ¬ЩҮвҖҢЫҢ Ш§Ш®Щ„Ш§ЩӮЫҢ ШіШ§ШҜЩҮ ЩҲ Щ…Ш«ШЁШӘ ШҜШұ ЩҫШ§ЫҢШ§ЩҶ ШҜШ§ШҙШӘЩҮ ШЁШ§ШҙШҜ\n"
        f"- ШҙШ®ШөЫҢШӘ Ш§ШөЩ„ЫҢ ШЁШ§ЫҢШҜ {gender_fa} ШЁШ§ШҙШҜ ШӘШ§ Ъ©ЩҲШҜЪ© ШЁЩҮШӘШұ ШЁШ§ ШўЩҶ Ш§ШұШӘШЁШ§Ш· ШЁЪҜЫҢШұШҜ\n"
        f"- Ш·ЩҲЩ„ ШҜШ§ШіШӘШ§ЩҶ Щ…ШӘЩҶШ§ШіШЁ ШЁШ§ ШіЩҶ Ш§ЩҶШӘШ®Ш§ШЁ ШҙЩҲШҜ: ШЁШұШ§ЫҢ ШіЩҶЫҢЩҶ Ъ©ЩҲЪҶЪ©вҖҢШӘШұ (ЫІ ШӘШ§ Ыө ШіШ§Щ„) Ъ©ЩҲШӘШ§ЩҮвҖҢШӘШұ "
        f"(ШҜШұ ШӯШҜ ЫІ ШӘШ§ Ыі ШөЩҒШӯЩҮШҢ ШӯШҜЩҲШҜ ЫөЫ°Ы° ШӘШ§ ЫёЫ°Ы° Ъ©Щ„Щ…ЩҮ) ЩҲ ШЁШұШ§ЫҢ ШіЩҶЫҢЩҶ ШЁШІШұЪҜвҖҢШӘШұ (Ы¶ ШӘШ§ ЫұЫҙ ШіШ§Щ„) "
        f"ШЁЩ„ЩҶШҜШӘШұ ЩҲ ШЁШ§ Ш¬ШІШҰЫҢШ§ШӘ ШЁЫҢШҙШӘШұ (ШӘШ§ Ы· ЫҢШ§ Ыё ШөЩҒШӯЩҮШҢ ШӯШҜЩҲШҜ ЫұЫёЫ°Ы° ШӘШ§ ЫІЫөЫ°Ы° Ъ©Щ„Щ…ЩҮ)Шӣ ШҜШұ ЩҮШұ ШӯШ§Щ„ШӘ "
        f"Ш·ЩҲЩ„ ЩҶЩҮШ§ЫҢЫҢ ШЁШ§ЫҢШҜ ШЁЫҢЩҶ ЫІ ШөЩҒШӯЩҮ ЩҲ Ыё ШөЩҒШӯЩҮ ШЁЩ…Ш§ЩҶШҜШҢ ЩҶЩҮ Ъ©Щ…ШӘШұ ЩҲ ЩҶЩҮ ШЁЫҢШҙШӘШұ\n"
        f"- ЩҒЩӮШ· Щ…ШӘЩҶ ШҜШ§ШіШӘШ§ЩҶ ШұШ§ ШЁЩҶЩҲЫҢШіШҢ ШЁШҜЩҲЩҶ Щ…ЩӮШҜЩ…ЩҮ ЫҢШ§ ШӘЩҲШ¶ЫҢШӯ Ш§Ш¶Ш§ЩҒЩҮ\n"
        f"- ШҜШұ ЩҫШ§ЫҢШ§ЩҶШҢ ЫҢЪ© Ш®Ш· ШЁШ§ Ш№ЩҶЩҲШ§ЩҶ В«рҹҢҹ ЩҫЫҢШ§Щ… ШҜШ§ШіШӘШ§ЩҶ:В» ЩҲ ЩҶШӘЫҢШ¬ЩҮвҖҢЫҢ Ш§Ш®Щ„Ш§ЩӮЫҢ Ъ©ЩҲШӘШ§ЩҮ Ш§Ш¶Ш§ЩҒЩҮ Ъ©ЩҶ"
    )

# в”Җв”Җв”Җ ШіШ§Ш®ШӘ PDF ЩҒШ§ШұШіЫҢ ШЁШұШ§ЫҢ ЩӮШөЩҮ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
FONT_CACHE_PATH = "/tmp/Vazirmatn-Regular.ttf"
FONT_DOWNLOAD_URLS = [
    "https://cdn.jsdelivr.net/gh/rastikerdar/vazirmatn@v33.003/fonts/ttf/Vazirmatn-Regular.ttf",
    "https://github.com/rastikerdar/vazirmatn/raw/master/fonts/ttf/Vazirmatn-Regular.ttf",
]

async def ensure_persian_font():
    """
    ШҜШ§ЩҶЩ„ЩҲШҜ ЩҲ Ъ©Шҙ Ъ©ШұШҜЩҶ ЩҒЩҲЩҶШӘ ЩҒШ§ШұШіЫҢ Vazirmatn ШҜШұ /tmp (ЫҢЪ©вҖҢШЁШ§Шұ ШҜШұ Ш·ЩҲЩ„ Ш§Ш¬ШұШ§ЫҢ ШұШЁШ§ШӘ).
    Ш§ЫҢЩҶ ШұЩҲШҙ ЩҶЫҢШ§ШІЫҢ ШЁЩҮ ШўЩҫЩ„ЩҲШҜ ШҜШіШӘЫҢ ЩҒШ§ЫҢЩ„ ЩҒЩҲЩҶШӘ ШҜШұ ЩҫШұЩҲЪҳЩҮ ЩҶШҜШ§ШұШҜ.
    """
    if os.path.exists(FONT_CACHE_PATH) and os.path.getsize(FONT_CACHE_PATH) > 10000:
        return FONT_CACHE_PATH
    for url in FONT_DOWNLOAD_URLS:
        try:
            async with httpx.AsyncClient(timeout=20, follow_redirects=True) as client:
                r = await client.get(url)
                if r.status_code == 200 and len(r.content) > 10000:
                    with open(FONT_CACHE_PATH, "wb") as f:
                        f.write(r.content)
                    print(f"вң… ЩҒЩҲЩҶШӘ ЩҒШ§ШұШіЫҢ ШЁШ§ Щ…ЩҲЩҒЩӮЫҢШӘ ШҜШ§ЩҶЩ„ЩҲШҜ ШҙШҜ Ш§ШІ: {url}")
                    return FONT_CACHE_PATH
        except Exception as e:
            print(f"Ш®Ш·Ш§ЫҢ ШҜШ§ЩҶЩ„ЩҲШҜ ЩҒЩҲЩҶШӘ Ш§ШІ {url}: {e}")
            continue
    print("вҡ пёҸ ЩҮЫҢЪҶвҖҢЪ©ШҜШ§Щ… Ш§ШІ Щ…ЩҶШ§ШЁШ№ ЩҒЩҲЩҶШӘ ЩҒШ§ШұШіЫҢ ШҜШұ ШҜШіШӘШұШі ЩҶШЁЩҲШҜЩҶШҜ.")
    return None

async def build_story_pdf(title, story_text):
    """
    ШіШ§Ш®ШӘ PDF ЩҒШ§ШұШіЫҢ ШЁШ§ fpdf2. ЩҒЩҲЩҶШӘ ЩҒШ§ШұШіЫҢ ШЁЩҮвҖҢШөЩҲШұШӘ Ш®ЩҲШҜЪ©Ш§Шұ Ш§ШІ CDN ШҜШ§ЩҶЩ„ЩҲШҜ ЩҲ Ъ©Шҙ Щ…ЫҢвҖҢШҙЩҲШҜ
    (ЩҶЫҢШ§ШІЫҢ ШЁЩҮ ШўЩҫЩ„ЩҲШҜ ШҜШіШӘЫҢ ЩҒШ§ЫҢЩ„ ЩҒЩҲЩҶШӘ ШҜШұ ЩҫШұЩҲЪҳЩҮ ЩҶЫҢШіШӘ).
    Ш§ЪҜШұ ШҜШ§ЩҶЩ„ЩҲШҜ ЩҒЩҲЩҶШӘ ЩҶШ§Щ…ЩҲЩҒЩӮ ШЁШ§ШҙШҜШҢ Ш®Ш·Ш§ ЪҶШ§Щҫ Щ…ЫҢвҖҢШҙЩҲШҜ ЩҲ ШӘШ§ШЁШ№ None ШЁШұЩ…ЫҢвҖҢЪҜШұШҜШ§ЩҶШҜ (ЫҢШ№ЩҶЫҢ ЩҒЩӮШ·
    Щ…ШӘЩҶ ШҜШұ ЪҶШӘ ЩҶЩ…Ш§ЫҢШҙ ШҜШ§ШҜЩҮ Щ…ЫҢвҖҢШҙЩҲШҜШҢ ШЁШҜЩҲЩҶ ЩҒШ§ЫҢЩ„ PDF).
    """
    try:
        from fpdf import FPDF
        font_path = await ensure_persian_font()
        if not font_path:
            return None

        pdf = FPDF()
        pdf.add_page()
        pdf.add_font("Vazirmatn", "", font_path, uni=True)
        pdf.set_font("Vazirmatn", size=16)
        pdf.set_text_shaping(True)  # Ш¶ШұЩҲШұЫҢ ШЁШұШ§ЫҢ ШұЩҶШҜШұ ШҜШұШіШӘ ШұШ§ШіШӘвҖҢШЁЩҮвҖҢЪҶЩҫ ЩҒШ§ШұШіЫҢ
        pdf.multi_cell(0, 12, title, align="C")
        pdf.ln(5)
        pdf.set_font("Vazirmatn", size=13)
        pdf.multi_cell(0, 10, story_text, align="R")

        buffer = io.BytesIO()
        pdf.output(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        print(f"Ш®Ш·Ш§ЫҢ ШіШ§Ш®ШӘ PDF ЩӮШөЩҮ: {e}")
        return None

# в”Җв”Җв”Җ ЩҮЩҶШҜЩ„Шұ ШҜЪ©Щ…ЩҮвҖҢЩҮШ§ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
async def handle_kids(query, user_id):
    data = query.data

    if data == "menu_kids":
        await query.edit_message_text("рҹ‘¶ ШЁШ®Шҙ ЩҒШұШІЩҶШҜШ§ЩҶ:", reply_markup=kids_menu())

    elif data == "kids_story":
        set_state(user_id, "kids_age")
        await query.edit_message_text(
            "рҹ“– ШіШ§Ш®ШӘ ЩӮШөЩҮ Ш¬ШҜЫҢШҜ\n\nШіЩҶ Ъ©ЩҲШҜЪ© ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ (Ш№ШҜШҜ):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="menu_kids")]])
        )

    elif data in ["kids_gender_girl", "kids_gender_boy"]:
        gender = "girl" if data == "kids_gender_girl" else "boy"
        state, sdata = get_state(user_id)
        if state == "kids_gender":
            set_state(user_id, "kids_topic", f"{sdata}|{gender}")
            await query.edit_message_text(
                "вңҸпёҸ Щ…ЩҲШ¶ЩҲШ№ ШҜШ§ШіШӘШ§ЩҶ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:\n\n"
                "Щ…Ш«Ш§Щ„: ШҜЩҲШіШӘЫҢШҢ ШҙШ¬Ш§Ш№ШӘШҢ ШөШҜШ§ЩӮШӘШҢ ШӯЫҢЩҲШ§ЩҶШ§ШӘШҢ ЩҒШ¶Ш§ШҢ Ш§ШӯШӘШұШ§Щ… ШЁЩҮ ШЁШІШұЪҜвҖҢШӘШұЩҮШ§"
            )

    elif data == "kids_homework":
        await query.answer("рҹ”§ ШЁЩҮ ШІЩҲШҜЫҢ Ш§Ш¶Ш§ЩҒЩҮ Щ…ЫҢвҖҢШҙЩҲШҜ!", show_alert=True)

async def handle_kids_message(user_id, text, update):
    state, data = get_state(user_id)
    text = text.strip()

    if state == "kids_age":
        try:
            age = int(text)
            if age < 2 or age > 14:
                await update.message.reply_text("вқҢ Щ„Ш·ЩҒШ§ЩӢ ШіЩҶЫҢ ШЁЫҢЩҶ ЫІ ШӘШ§ ЫұЫҙ ШіШ§Щ„ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
                return True
            set_state(user_id, "kids_gender", str(age))
            await update.message.reply_text("Ш¬ЩҶШіЫҢШӘ Ъ©ЩҲШҜЪ© ШұШ§ Ш§ЩҶШӘШ®Ш§ШЁ Ъ©ЩҶЫҢШҜ:", reply_markup=kids_gender_menu())
        except:
            await update.message.reply_text("вқҢ Щ„Ш·ЩҒШ§ЩӢ ЫҢЪ© Ш№ШҜШҜ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        return True

    elif state == "kids_topic":
        topic = text
        parts = data.split("|")
        age, gender = parts[0], parts[1]

        await update.message.reply_text("вҸі ШҜШұ ШӯШ§Щ„ ЩҶЩҲШҙШӘЩҶ ЩӮШөЩҮ... (Щ…Щ…Ъ©ЩҶ Ш§ШіШӘ Ъ©Щ…ЫҢ Ш·ЩҲЩ„ ШЁЪ©ШҙШҜ)")

        prompt = build_story_prompt(age, gender, topic)
        messages = [
            {"role": "system", "content": "ШӘЩҲ ЫҢЪ© ЩҶЩҲЫҢШіЩҶШҜЩҮвҖҢЫҢ Ш®Щ„Ш§ЩӮ ШҜШ§ШіШӘШ§ЩҶвҖҢЩҮШ§ЫҢ Ъ©ЩҲШҜЪ©Ш§ЩҶЩҮ ШЁЩҮ ШІШЁШ§ЩҶ ЩҒШ§ШұШіЫҢ ЩҮШіШӘЫҢ."},
            {"role": "user", "content": prompt}
        ]
        story = await get_ai_response(messages)

        if not story:
            await update.message.reply_text(
                "вқҢ Щ…ШӘШЈШіЩҒШ§ЩҶЩҮ ШіШ§Ш®ШӘ ЩӮШөЩҮ Щ…ЩҲЩӮШӘШ§ЩӢ Щ…Щ…Ъ©ЩҶ ЩҶЫҢШіШӘ. Ъ©Щ…ЫҢ ШЁШ№ШҜ ШҜЩҲШЁШ§ШұЩҮ Ш§Щ…ШӘШӯШ§ЩҶ Ъ©ЩҶЫҢШҜ.",
                reply_markup=kids_menu()
            )
            clear_state(user_id)
            return True

        title = f"ЩӮШөЩҮвҖҢШ§ЫҢ ШҜШұШЁШ§ШұЩҮ {topic}"
        full_text = f"рҹ“– {title}\n\n{story}"

        # ШӘЩ„ЪҜШұШ§Щ… ЩҮШұ ЩҫЫҢШ§Щ… ШұШ§ ШӯШҜШ§Ъ©Ш«Шұ ШӘШ§ ЫҙЫ°Ы№Ы¶ Ъ©Ш§ШұШ§Ъ©ШӘШұ Щ…ЫҢвҖҢЩҫШ°ЫҢШұШҜШӣ ЩӮШөЩҮвҖҢЩҮШ§ЫҢ ШЁЩ„ЩҶШҜШӘШұ ШӘЪ©ЩҮвҖҢШӘЪ©ЩҮ Ш§ШұШіШ§Щ„ Щ…ЫҢвҖҢШҙЩҲЩҶШҜ
        TELEGRAM_LIMIT = 4000
        if len(full_text) <= TELEGRAM_LIMIT:
            await update.message.reply_text(full_text, reply_markup=kids_menu())
        else:
            chunks = [full_text[i:i + TELEGRAM_LIMIT] for i in range(0, len(full_text), TELEGRAM_LIMIT)]
            for i, chunk in enumerate(chunks):
                is_last = (i == len(chunks) - 1)
                await update.message.reply_text(chunk, reply_markup=kids_menu() if is_last else None)

        pdf_buffer = await build_story_pdf(title, story)
        if pdf_buffer:
            await update.message.reply_document(
                document=pdf_buffer,
                filename=f"{title}.pdf",
                caption="рҹ“Ҙ ЩҶШіШ®ЩҮ PDF ЩӮШөЩҮ ШЁШұШ§ЫҢ ЩҶЪҜЩҮШҜШ§ШұЫҢ ЫҢШ§ ЪҶШ§Щҫ"
            )

        clear_state(user_id)
        return True

    return False

# в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ ЩҫШ§ЫҢШ§ЩҶ MODULE: KIDS в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ




# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                  MODULE: FUN (ШіШұЪҜШұЩ…ЫҢ)                        в•‘
# в•‘  Ш§ЫҢЩҶ Щ…Ш§ЪҳЩҲЩ„ ШҙШ§Щ…Щ„ ШіЩҮ ШІЫҢШұШЁШ®Шҙ Щ…ШіШӘЩӮЩ„ Ш§ШіШӘ:                        в•‘
# в•‘    - ШіЩ„Ш§Щ…ШӘ ЩҲ ЩҲШұШІШҙ (Щ…ЩҶШӘЩӮЩ„вҖҢШҙШҜЩҮ Ш§ШІ Щ…ЩҶЩҲЫҢ Ш§ШөЩ„ЫҢ)                  в•‘
# в•‘    - Ъ©ШӘШ§ШЁвҖҢШ®ЩҲШ§ЩҶЫҢ                                              в•‘
# в•‘    - Щ…ЩҲШІЫҢЪ© (Ш¬ШіШӘШ¬ЩҲ ШҜШұ Ъ©Ш§ЩҶШ§Щ„ + Щ…ЩҶШ§ШЁШ№ ШұШ§ЫҢЪҜШ§ЩҶ ЩӮШ§ЩҶЩҲЩҶЫҢ)           в•‘
# в•‘  ШЁШұШ§ЫҢ ШӯШ°ЩҒ Ъ©Ш§Щ…Щ„: Ш§ЫҢЩҶ ШЁЩ„ЩҲЪ© ШұЩҲ ЩҫШ§Ъ© Ъ©ЩҶ                          в•‘
# в•‘  + Ш®Ш· menu_fun ШұЩҲ Ш§ШІ main_menu ЩҫШ§Ъ© Ъ©ЩҶ                       в•‘
# в•‘  + Ш®Ш· fun_/health_/book_/music_ ШұЩҲ Ш§ШІ ROUTER ЩҫШ§Ъ© Ъ©ЩҶ         в•‘
# в•‘  + state ЩҮШ§ЫҢ fun_/book_/music_ ШұЩҲ Ш§ШІ message_handler ЩҫШ§Ъ© Ъ©ЩҶв•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ

MUSIC_CHANNEL_USERNAME = "LiteMusics"  # ШЁШҜЩҲЩҶ @ ЩҲ ШЁШҜЩҲЩҶ t.me/

# в”Җв”Җв”Җ ШҜЫҢШӘШ§ШЁЫҢШі Ш§ЫҢЩҶ Щ…Ш§ЪҳЩҲЩ„ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
def init_fun_db():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS book_tracking (
        user_id     INTEGER PRIMARY KEY,
        title       TEXT,
        current_page INTEGER DEFAULT 0,
        total_pages  INTEGER,
        updated_at   TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS music_index (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        artist      TEXT,
        title       TEXT,
        file_id     TEXT,
        message_id  INTEGER,
        added_at    TEXT
    )''')
    conn.commit()
    conn.close()

def save_book_progress(user_id, title, current_page, total_pages):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''INSERT INTO book_tracking (user_id, title, current_page, total_pages, updated_at)
                 VALUES (?, ?, ?, ?, ?)
                 ON CONFLICT(user_id) DO UPDATE SET
                    title=excluded.title, current_page=excluded.current_page,
                    total_pages=excluded.total_pages, updated_at=excluded.updated_at''',
              (user_id, title, current_page, total_pages, datetime.now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def get_book_progress(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT title, current_page, total_pages, updated_at FROM book_tracking WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row

def update_book_page(user_id, new_page):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE book_tracking SET current_page=?, updated_at=? WHERE user_id=?",
              (new_page, datetime.now().strftime("%Y-%m-%d %H:%M"), user_id))
    conn.commit()
    conn.close()

def index_music_track(artist, title, file_id, message_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    # Ш§ШІ ШӘЪ©ШұШ§Шұ Ш¬Щ„ЩҲЪҜЫҢШұЫҢ ШҙЩҲШҜ (ЩҮЩ…Ш§ЩҶ message_id ШҜЩҲШЁШ§ШұЩҮ Ш§ЫҢЩҶШҜЪ©Ші ЩҶШҙЩҲШҜ)
    c.execute("SELECT id FROM music_index WHERE message_id=?", (message_id,))
    if c.fetchone():
        conn.close()
        return
    c.execute("INSERT INTO music_index (artist, title, file_id, message_id, added_at) VALUES (?, ?, ?, ?, ?)",
              (artist, title, file_id, message_id, datetime.now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def search_music_index(query_text):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    like = f"%{query_text}%"
    c.execute("SELECT artist, title, file_id FROM music_index WHERE artist LIKE ? OR title LIKE ? ORDER BY id DESC LIMIT 5",
              (like, like))
    rows = c.fetchall()
    conn.close()
    return rows

# в”Җв”Җв”Җ ЩҫШ§ШұШі Ъ©ЩҫШҙЩҶ ЩҫШіШӘ Ъ©Ш§ЩҶШ§Щ„ ШЁШұШ§ЫҢ Ш§ШіШӘШ®ШұШ§Ш¬ Ш®ЩҲШ§ЩҶЩҶШҜЩҮ/Ш§ШіЩ… ШўЩҮЩҶЪҜ в”Җв”Җв”Җв”Җв”Җв”Җ
def parse_music_caption(caption):
    """
    ЩҒШұЩ…ШӘ ЩҫШіШӘвҖҢЩҮШ§ЫҢ Ъ©Ш§ЩҶШ§Щ„:
    рҹҺӨ  #ЩҶШ§Щ…_Ш®ЩҲШ§ЩҶЩҶШҜЩҮ
    рҹҺј    ЩҶШ§Щ… ШўЩҮЩҶЪҜ
    [ШЁЩӮЫҢЩҮ Щ…ШӘЩҶ...]
    """
    if not caption:
        return None, None
    artist, title = None, None
    for line in caption.split("\n"):
        line = line.strip()
        if line.startswith("рҹҺӨ"):
            artist = line.replace("рҹҺӨ", "").replace("#", "").strip()
        elif line.startswith("рҹҺј"):
            title = line.replace("рҹҺј", "").strip()
    return artist, title

# в”Җв”Җв”Җ Ш¬ШіШӘШ¬ЩҲ ШҜШұ Щ…ЩҶШ§ШЁШ№ ШұШ§ЫҢЪҜШ§ЩҶ ЩҲ ЩӮШ§ЩҶЩҲЩҶЫҢ (Jamendo API) в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
# Jamendo: Ъ©Ш§ШӘШ§Щ„ЩҲЪҜ +ЫөЫ°Ы°,Ы°Ы°Ы° ШӘШұЪ© ШЁШ§ Щ…Ш¬ЩҲШІ Creative CommonsШҢ ШұШ§ЫҢЪҜШ§ЩҶ
# ШЁШұШ§ЫҢ ЩҒШ№Ш§Щ„вҖҢШіШ§ШІЫҢ Ъ©Ш§Щ…Щ„: ЫҢЪ© client_id ШұШ§ЫҢЪҜШ§ЩҶ Ш§ШІ https://devportal.jamendo.com ШЁЪҜЫҢШұЫҢШҜ
# ЩҲ ШҜШұ Щ…ШӘШәЫҢШұ Щ…ШӯЫҢШ·ЫҢ JAMENDO_CLIENT_ID ЩӮШұШ§Шұ ШҜЩҮЫҢШҜ. ШЁШҜЩҲЩҶ ШўЩҶШҢ Ш§ЫҢЩҶ ШЁШ®Шҙ ЩҶШӘЫҢШ¬ЩҮвҖҢШ§ЫҢ ШЁШұЩҶЩ…ЫҢвҖҢЪҜШұШҜШ§ЩҶШҜ
# ЩҲ Щ…ШіШӘЩӮЫҢЩ…Ш§ЩӢ ШЁЩҮ Щ…ШұШӯЩ„ЩҮ В«Ш§Ш·Щ„Ш§Ш№ ШЁЩҮ Ш§ШҜЩ…ЫҢЩҶВ» Щ…ЫҢвҖҢШұЩҲШҜ (Ъ©ЩҮ ЩҮЩ…ЪҶЩҶШ§ЩҶ Ъ©Ш§Шұ Щ…ЫҢвҖҢЪ©ЩҶШҜ).
JAMENDO_CLIENT_ID = os.environ.get("JAMENDO_CLIENT_ID", "")

async def search_free_legal_music(query_text):
    """
    Ш¬ШіШӘШ¬ЩҲ ШҜШұ Jamendo - ЩҒЩӮШ· Щ…ЩҲШІЫҢЪ©вҖҢЩҮШ§ЫҢЫҢ ШЁШ§ Щ…Ш¬ЩҲШІ Creative Commons (Ъ©Ш§Щ…Щ„Ш§ЩӢ ЩӮШ§ЩҶЩҲЩҶЫҢ)
    Ш®ШұЩҲШ¬ЫҢ: (title, artist, download_url) ЫҢШ§ None
    """
    if not JAMENDO_CLIENT_ID:
        return None
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(
                "https://api.jamendo.com/v3.0/tracks/",
                params={
                    "client_id": JAMENDO_CLIENT_ID,
                    "format": "json",
                    "limit": 1,
                    "namesearch": query_text,
                }
            )
            data = r.json()
            results = data.get("results", [])
            if not results:
                return None
            track = results[0]
            title = track.get("name", "ЩҶШ§Щ…ШҙШ®Шө")
            artist = track.get("artist_name", "ЩҶШ§Щ…ШҙШ®Шө")
            download_url = track.get("audiodownload") or track.get("audio")
            if not download_url:
                return None
            return title, artist, download_url
    except Exception as e:
        print(f"Ш®Ш·Ш§ЫҢ Ш¬ШіШӘШ¬ЩҲЫҢ Щ…ЩҲШІЫҢЪ© ШұШ§ЫҢЪҜШ§ЩҶ: {e}")
        return None

# в”Җв”Җв”Җ Щ…ЩҶЩҲЩҮШ§ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
def fun_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹҸӢпёҸ ШіЩ„Ш§Щ…ШӘ ЩҲ ЩҲШұШІШҙ", callback_data="menu_health")],
        [InlineKeyboardButton("рҹ“– Ъ©ШӘШ§ШЁвҖҢШ®ЩҲШ§ЩҶЫҢ", callback_data="menu_book")],
        [InlineKeyboardButton("рҹҺө Щ…ЩҲШІЫҢЪ©", callback_data="menu_music")],
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_main")],
    ])

def health_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("вҡ–пёҸ Ш«ШЁШӘ ЩҲШІЩҶ", callback_data="health_weight")],
        [InlineKeyboardButton("рҹ“Ҡ ЩҶЩ…ЩҲШҜШ§Шұ ЩҲШІЩҶ", callback_data="health_chart")],
        [InlineKeyboardButton("рҹҸғ ШЁШұЩҶШ§Щ…ЩҮ ЩҲШұШІШҙЫҢ", callback_data="health_exercise")],
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_fun")],
    ])

def book_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("вһ• Ш«ШЁШӘ Ъ©ШӘШ§ШЁ Ш¬ШҜЫҢШҜ", callback_data="book_new")],
        [InlineKeyboardButton("рҹ“Ҡ ЩҫЫҢЪҜЫҢШұЫҢ Ъ©ШӘШ§ШЁ ЩҒШ№Щ„ЫҢ", callback_data="book_progress")],
        [InlineKeyboardButton("рҹ’Ў Щ…Ш№ШұЩҒЫҢ Ъ©ШӘШ§ШЁ Ш¬ШҜЫҢШҜ", callback_data="book_suggest")],
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_fun")],
    ])

def music_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹ”Қ Ш¬ШіШӘШ¬ЩҲЫҢ ШўЩҮЩҶЪҜ", callback_data="music_search")],
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_fun")],
    ])

BOOK_SUGGESTIONS = [
    ("Щ…Щ„ШӘ Ш№ШҙЩӮ", "Ш§Щ„ЫҢЩҒ ШҙШ§ЩҒШ§Ъ©", "Ш№Ш§ШҙЩӮШ§ЩҶЩҮ/ЩҒЩ„ШіЩҒЫҢ"),
    ("Ъ©ЫҢЩ…ЫҢШ§ЪҜШұ", "ЩҫШ§ШҰЩҲЩ„ЩҲ Ъ©ЩҲШҰЫҢЩ„ЩҲ", "ЩҒЩ„ШіЩҒЫҢ/Ш§Щ„ЩҮШ§Щ…вҖҢШЁШ®Шҙ"),
    ("ШөШҜ ШіШ§Щ„ ШӘЩҶЩҮШ§ЫҢЫҢ", "ЪҜШ§ШЁШұЫҢЩ„ ЪҜШ§ШұШіЫҢШ§ Щ…Ш§ШұЪ©ШІ", "ШұШҰШ§Щ„ЫҢШіЩ… Ш¬Ш§ШҜЩҲЫҢЫҢ"),
    ("ШЁЩҲЩҒ Ъ©ЩҲШұ", "ШөШ§ШҜЩӮ ЩҮШҜШ§ЫҢШӘ", "Ш§ШҜШЁЫҢШ§ШӘ Ъ©Щ„Ш§ШіЫҢЪ© ЩҒШ§ШұШіЫҢ"),
    ("Ш¬ЩҶШ§ЫҢШӘ ЩҲ Щ…Ъ©Ш§ЩҒШ§ШӘ", "ШҜШ§ШіШӘШ§ЫҢЩҲЩҒШіЪ©ЫҢ", "ШұЩҲШ§ЩҶвҖҢШҙЩҶШ§Ш®ШӘЫҢ/Ъ©Щ„Ш§ШіЫҢЪ©"),
    ("Ъ©Ш§ЩҒЪ©Ш§ ШҜШұ Ъ©ЩҶШ§Шұ ШҜШұЫҢШ§", "ЩҮШ§ШұЩҲЪ©ЫҢ Щ…ЩҲШұШ§Ъ©Ш§Щ…ЫҢ", "ШұШҰШ§Щ„ЫҢШіЩ… Ш¬Ш§ШҜЩҲЫҢЫҢ Щ…ШҜШұЩҶ"),
    ("ШҙШ§ШІШҜЩҮ Ъ©ЩҲЪҶЩҲЩ„ЩҲ", "ШўЩҶШӘЩҲШ§ЩҶ ШҜЩҲ ШіЩҶШӘ Ш§ЪҜШІЩҲЩҫШұЫҢ", "ЩҒЩ„ШіЩҒЫҢ/Ъ©ЩҲШҜЪ© ЩҲ ШЁШІШұЪҜШіШ§Щ„"),
    ("ЩҮШІШ§Шұ Ш®ЩҲШұШҙЫҢШҜ ШӘШ§ШЁШ§ЩҶ", "Ш®Ш§Щ„ШҜ ШӯШіЫҢЩҶЫҢ", "ШҜШұШ§Щ… Ш§Ш¬ШӘЩ…Ш§Ш№ЫҢ"),
]

# в”Җв”Җв”Җ ЩҮЩҶШҜЩ„Шұ ШҜЪ©Щ…ЩҮвҖҢЩҮШ§ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
async def handle_fun(query, user_id):
    data = query.data

    if data == "menu_fun":
        await query.edit_message_text("рҹҺ® ШіШұЪҜШұЩ…ЫҢ:", reply_markup=fun_menu())

    elif data == "back_fun":
        clear_state(user_id)
        await query.edit_message_text("рҹҺ® ШіШұЪҜШұЩ…ЫҢ:", reply_markup=fun_menu())

    # в”Җв”Җв”Җ ШіЩ„Ш§Щ…ШӘ ЩҲ ЩҲШұШІШҙ (ЩҮЩ…Ш§ЩҶ Щ…ЩҶШ·ЩӮ ЩӮШЁЩ„ЫҢ) в”Җв”Җв”Җ
    elif data == "menu_health":
        await query.edit_message_text("рҹҸӢпёҸ ШіЩ„Ш§Щ…ШӘ ЩҲ ЩҲШұШІШҙ:", reply_markup=health_menu())

    elif data.startswith("health_"):
        await query.answer("рҹ”§ ШЁЩҮ ШІЩҲШҜЫҢ Ш§Ш¶Ш§ЩҒЩҮ Щ…ЫҢвҖҢШҙЩҲШҜ!", show_alert=True)

    # в”Җв”Җв”Җ Ъ©ШӘШ§ШЁвҖҢШ®ЩҲШ§ЩҶЫҢ в”Җв”Җв”Җ
    elif data == "menu_book":
        await query.edit_message_text("рҹ“– Ъ©ШӘШ§ШЁвҖҢШ®ЩҲШ§ЩҶЫҢ:", reply_markup=book_menu())

    elif data == "book_new":
        set_state(user_id, "book_title")
        await query.edit_message_text(
            "рҹ“– ЩҶШ§Щ… Ъ©ШӘШ§ШЁЫҢ Ъ©ЩҮ ШҙШұЩҲШ№ ШЁЩҮ Ш®ЩҲШ§ЩҶШҜЩҶШҙ Щ…ЫҢвҖҢЪ©ЩҶЫҢШҜ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="menu_book")]])
        )

    elif data == "book_progress":
        progress = get_book_progress(user_id)
        if not progress:
            await query.edit_message_text(
                "рҹ“ӯ ЩҮЩҶЩҲШІ Ъ©ШӘШ§ШЁЫҢ Ш«ШЁШӘ ЩҶЪ©ШұШҜЩҮвҖҢШ§ЫҢШҜ.",
                reply_markup=book_menu()
            )
            return
        title, current_page, total_pages, updated_at = progress
        percent = (current_page / total_pages * 100) if total_pages else 0
        bar_filled = int(percent // 10)
        bar = "рҹҹ©" * bar_filled + "в¬ңпёҸ" * (10 - bar_filled)
        text = (
            f"рҹ“– {title}\n"
            f"{bar} {percent:.0f}%\n\n"
            f"ШөЩҒШӯЩҮ ЩҒШ№Щ„ЫҢ: {current_page} Ш§ШІ {total_pages}\n"
            f"ШўШ®ШұЫҢЩҶ ШЁЩҮвҖҢШұЩҲШІШұШіШ§ЩҶЫҢ: {updated_at}"
        )
        await query.edit_message_text(
            text,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("вңҸпёҸ ШЁЩҮвҖҢШұЩҲШІШұШіШ§ЩҶЫҢ ШөЩҒШӯЩҮ", callback_data="book_update_page")],
                [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="menu_book")],
            ])
        )

    elif data == "book_update_page":
        progress = get_book_progress(user_id)
        if not progress:
            await query.answer("Ш§ШЁШӘШҜШ§ ЫҢЪ© Ъ©ШӘШ§ШЁ Ш«ШЁШӘ Ъ©ЩҶЫҢШҜ.", show_alert=True)
            return
        set_state(user_id, "book_new_page")
        await query.edit_message_text("рҹ“„ ШҙЩ…Ш§ШұЩҮ ШөЩҒШӯЩҮ ЩҒШ№Щ„ЫҢ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")

    elif data == "book_suggest":
        import random
        title, author, genre = random.choice(BOOK_SUGGESTIONS)
        await query.edit_message_text(
            f"рҹ’Ў ЩҫЫҢШҙЩҶЩҮШ§ШҜ Ъ©ШӘШ§ШЁ\n\n"
            f"рҹ“• {title}\n"
            f"вңҚпёҸ ЩҶЩҲЫҢШіЩҶШҜЩҮ: {author}\n"
            f"рҹҸ· ЪҳШ§ЩҶШұ: {genre}",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("рҹ”„ ЩҫЫҢШҙЩҶЩҮШ§ШҜ ШҜЫҢЪҜШұ", callback_data="book_suggest")],
                [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="menu_book")],
            ])
        )

    # в”Җв”Җв”Җ Щ…ЩҲШІЫҢЪ© в”Җв”Җв”Җ
    elif data == "menu_music":
        await query.edit_message_text(
            f"рҹҺө Щ…ЩҲШІЫҢЪ©\n\n"
            f"ЩҶШ§Щ… ШўЩҮЩҶЪҜ ЫҢШ§ Ш®ЩҲШ§ЩҶЩҶШҜЩҮ ШұШ§ ШӘШ§ЫҢЩҫ Ъ©ЩҶЫҢШҜ ШӘШ§ Ш¬ШіШӘШ¬ЩҲ Ъ©ЩҶЩ….\n\n"
            f"Ш§ЩҲЩ„ ШҜШұ Ъ©Ш§ЩҶШ§Щ„ Щ…ЩҲШІЫҢЪ© Ш¬ШіШӘШ¬ЩҲ Щ…ЫҢвҖҢЪ©ЩҶЩ…ШҢ Ш§ЪҜШұ ЩҶШЁЩҲШҜ ШіШұШ§Шә "
            f"Щ…ЩҶШ§ШЁШ№ ШұШ§ЫҢЪҜШ§ЩҶ ЩҲ ЩӮШ§ЩҶЩҲЩҶЫҢ Щ…ЫҢвҖҢШұЩҲЩ….",
            reply_markup=music_menu()
        )

    elif data == "music_search":
        set_state(user_id, "music_query")
        await query.edit_message_text(
            "рҹ”Қ ЩҶШ§Щ… ШўЩҮЩҶЪҜ ЫҢШ§ Ш®ЩҲШ§ЩҶЩҶШҜЩҮ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="menu_music")]])
        )

async def handle_fun_message(user_id, text, update):
    state, data = get_state(user_id)
    text = text.strip()

    if state == "book_title":
        set_state(user_id, "book_total_pages", text)
        await update.message.reply_text("рҹ“„ ШӘШ№ШҜШ§ШҜ Ъ©Щ„ ШөЩҒШӯШ§ШӘ Ъ©ШӘШ§ШЁ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        return True

    elif state == "book_total_pages":
        try:
            total_pages = int(text)
            title = data
            save_book_progress(user_id, title, 0, total_pages)
            await update.message.reply_text(
                f"вң… Ъ©ШӘШ§ШЁ В«{title}В» Ш«ШЁШӘ ШҙШҜ!\nЩҮШұ ЩҲЩӮШӘ ЩҫЫҢШҙШұЩҒШӘ ШҜШ§ШҙШӘЫҢШҜШҢ Ш§ШІ Щ…ЩҶЩҲЫҢ Ъ©ШӘШ§ШЁвҖҢШ®ЩҲШ§ЩҶЫҢ ШЁЩҮвҖҢШұЩҲШІШұШіШ§ЩҶЫҢ Ъ©ЩҶЫҢШҜ.",
                reply_markup=book_menu()
            )
        except:
            await update.message.reply_text("вқҢ Щ„Ш·ЩҒШ§ЩӢ ЫҢЪ© Ш№ШҜШҜ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        return True

    elif state == "book_new_page":
        try:
            new_page = int(text)
            progress = get_book_progress(user_id)
            if progress and new_page > progress[2]:
                await update.message.reply_text(
                    f"вқҢ ШҙЩ…Ш§ШұЩҮ ШөЩҒШӯЩҮ ЩҶЩ…ЫҢвҖҢШӘЩҲШ§ЩҶШҜ ШЁЫҢШҙШӘШұ Ш§ШІ ШӘШ№ШҜШ§ШҜ Ъ©Щ„ ШөЩҒШӯШ§ШӘ ({progress[2]}) ШЁШ§ШҙШҜ:"
                )
                return True
            update_book_page(user_id, new_page)
            await update.message.reply_text("вң… ЩҫЫҢШҙШұЩҒШӘ ШЁЩҮвҖҢШұЩҲШІШұШіШ§ЩҶЫҢ ШҙШҜ!", reply_markup=book_menu())
        except:
            await update.message.reply_text("вқҢ Щ„Ш·ЩҒШ§ЩӢ ЫҢЪ© Ш№ШҜШҜ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        return True

    elif state == "music_query":
        query_text = text
        await update.message.reply_text("вҸі ШҜШұ ШӯШ§Щ„ Ш¬ШіШӘШ¬ЩҲ...")

        # Ыұ) Ш¬ШіШӘШ¬ЩҲ ШҜШұ Ш§ЫҢЩҶШҜЪ©Ші Ъ©Ш§ЩҶШ§Щ„
        results = search_music_index(query_text)
        if results:
            artist, title, file_id = results[0]
            await update.message.reply_audio(
                audio=file_id,
                caption=f"рҹҺӨ {artist}\nрҹҺј {title}\n\nШ§ШІ Ъ©Ш§ЩҶШ§Щ„: @{MUSIC_CHANNEL_USERNAME}"
            )
            clear_state(user_id)
            await update.message.reply_text("ЩҶШӘЫҢШ¬ЩҮ ЩҫЫҢШҜШ§ ШҙШҜ вң…", reply_markup=music_menu())
            return True

        # ЫІ) Ш¬ШіШӘШ¬ЩҲ ШҜШұ Щ…ЩҶШ§ШЁШ№ ШұШ§ЫҢЪҜШ§ЩҶ ЩҲ ЩӮШ§ЩҶЩҲЩҶЫҢ
        free_result = await search_free_legal_music(query_text)
        if free_result:
            title, artist, url = free_result
            await update.message.reply_text(
                f"рҹҺө ЩҫЫҢШҜШ§ ШҙШҜ ШҜШұ Щ…ЩҶШЁШ№ ШұШ§ЫҢЪҜШ§ЩҶ ЩҲ ЩӮШ§ЩҶЩҲЩҶЫҢ:\n\n"
                f"рҹҺј {title}\nрҹҺӨ {artist}\n\nрҹ”— Щ„ЫҢЩҶЪ© ШҜШ§ЩҶЩ„ЩҲШҜ:\n{url}",
                reply_markup=music_menu()
            )
            clear_state(user_id)
            return True

        # Ыі) ЩҫЫҢШҜШ§ ЩҶШҙШҜ - Ш§Ш·Щ„Ш§Ш№ ШЁЩҮ Ш§ШҜЩ…ЫҢЩҶ
        await update.message.reply_text(
            f"вқҢ Щ…ШӘШЈШіЩҒШ§ЩҶЩҮ В«{query_text}В» ШҜШұ Ъ©Ш§ЩҶШ§Щ„ ЫҢШ§ Щ…ЩҶШ§ШЁШ№ ШұШ§ЫҢЪҜШ§ЩҶ ЩҫЫҢШҜШ§ ЩҶШҙШҜ.\n\n"
            f"ШҜШұШ®ЩҲШ§ШіШӘ ШҙЩ…Ш§ ШЁШұШ§ЫҢ Ш§ШҜЩ…ЫҢЩҶ Ш§ШұШіШ§Щ„ ШҙШҜ ШӘШ§ ШҜШұ ШөЩҲШұШӘ ШҜШ§ШҙШӘЩҶ Щ…Ш¬ЩҲШІ ЩҫШ®ШҙШҢ Ш§Ш¶Ш§ЩҒЩҮ ШҙЩҲШҜ.",
            reply_markup=music_menu()
        )
        try:
            requester = update.effective_user
            await update.get_bot().send_message(
                ADMIN_ID,
                f"рҹҺө ШҜШұШ®ЩҲШ§ШіШӘ ШўЩҮЩҶЪҜ ШЁШҜЩҲЩҶ ЩҶШӘЫҢШ¬ЩҮ:\n\n"
                f"рҹ”Қ Ш¬ШіШӘШ¬ЩҲ: {query_text}\n"
                f"рҹ‘Ө Ш§ШІ Ш·ШұЩҒ: {requester.first_name} (@{requester.username or 'ЩҶШҜШ§ШұШҜ'})\n\n"
                f"ШҜШұ ШөЩҲШұШӘ ШҜШ§ШҙШӘЩҶ Щ…Ш¬ЩҲШІ ЩҫШ®ШҙШҢ Щ„Ш·ЩҒШ§ЩӢ ШЁЩҮ Ъ©Ш§ЩҶШ§Щ„ @{MUSIC_CHANNEL_USERNAME} Ш§Ш¶Ш§ЩҒЩҮ Ъ©ЩҶЫҢШҜ."
            )
        except Exception as e:
            print(f"Ш®Ш·Ш§ ШҜШұ Ш§Ш·Щ„Ш§Ш№ ШЁЩҮ Ш§ШҜЩ…ЫҢЩҶ: {e}")
        clear_state(user_id)
        return True

    return False

# в”Җв”Җв”Җ Ш§ЫҢЩҶШҜЪ©Ші Ъ©ШұШҜЩҶ ЩҫШіШӘвҖҢЩҮШ§ЫҢ Ш¬ШҜЫҢШҜ Ъ©Ш§ЩҶШ§Щ„ Щ…ЩҲШІЫҢЪ© (Ш®ЩҲШҜЪ©Ш§Шұ) в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
async def handle_channel_post(update: Update, context: ContextTypes.DEFAULT_TYPE):
    post = update.channel_post
    if not post:
        return
    chat_username = post.chat.username
    if not chat_username or chat_username.lower() != MUSIC_CHANNEL_USERNAME.lower():
        return
    if not post.audio:
        return
    caption = post.caption or ""
    artist, title = parse_music_caption(caption)
    if not artist:
        artist = post.audio.performer or "ЩҶШ§Щ…ШҙШ®Шө"
    if not title:
        title = post.audio.title or "ЩҶШ§Щ…ШҙШ®Шө"
    index_music_track(artist, title, post.audio.file_id, post.message_id)
    print(f"рҹҺө ШўЩҮЩҶЪҜ Ш¬ШҜЫҢШҜ Ш§ЫҢЩҶШҜЪ©Ші ШҙШҜ: {artist} - {title}")

# в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ ЩҫШ§ЫҢШ§ЩҶ MODULE: FUN в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ





# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                  MODULE: LEARNING                            в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ

def learn_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹ”Ө Ъ©Щ„Щ…ЩҮ Ш§ЩҶЪҜЩ„ЫҢШіЫҢ ШұЩҲШІШ§ЩҶЩҮ", callback_data="learn_word")],
        [InlineKeyboardButton("рҹ“ҳ Ш®Щ„Ш§ШөЩҮ Ъ©ШӘШ§ШЁ", callback_data="learn_book")],
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_main")],
    ])

async def handle_learning(query, user_id):
    data = query.data
    if data == "menu_learn":
        await query.edit_message_text("рҹ“ҡ ЫҢШ§ШҜЪҜЫҢШұЫҢ:", reply_markup=learn_menu())
    else:
        await query.answer("рҹ”§ ШЁЩҮ ШІЩҲШҜЫҢ Ш§Ш¶Ш§ЩҒЩҮ Щ…ЫҢвҖҢШҙЩҲШҜ!", show_alert=True)

# в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ ЩҫШ§ЫҢШ§ЩҶ MODULE: LEARNING в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ


# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                  MODULE: SETTINGS                            в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ

def settings_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹ‘Ө ЩҫШұЩҲЩҒШ§ЫҢЩ„ Щ…ЩҶ", callback_data="set_profile")],
        [InlineKeyboardButton("рҹ“… ШӘШ§ШұЫҢШ® Ш§ЩҶЩӮШ¶Ш§", callback_data="set_expiry")],
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_main")],
    ])

async def handle_settings(query, user_id):
    data = query.data
    if data == "menu_settings":
        await query.edit_message_text("вҡҷпёҸ ШӘЩҶШёЫҢЩ…Ш§ШӘ:", reply_markup=settings_menu())
    elif data == "set_profile":
        db_user = get_user(user_id)
        expires = db_user[5] if db_user else "---"
        await query.edit_message_text(
            f"рҹ‘Ө ЩҫШұЩҲЩҒШ§ЫҢЩ„ Щ…ЩҶ\n\nрҹҶ” ШўЫҢШҜЫҢ: {user_id}\nрҹ“… Ш§ЩҶЩӮШ¶Ш§: {expires}",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="menu_settings")]]))
    elif data == "set_expiry":
        db_user = get_user(user_id)
        expires = db_user[5] if db_user else "---"
        await query.answer(f"рҹ“… Ш§ЩҶЩӮШ¶Ш§: {expires}", show_alert=True)
    else:
        await query.answer("рҹ”§ ШЁЩҮ ШІЩҲШҜЫҢ Ш§Ш¶Ш§ЩҒЩҮ Щ…ЫҢвҖҢШҙЩҲШҜ!", show_alert=True)

# в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ ЩҫШ§ЫҢШ§ЩҶ MODULE: SETTINGS в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ

# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                  MODULE: TRANSLATE                           в•‘
# в•‘  ШЁШұШ§ЫҢ ШӯШ°ЩҒ: Ш§ЫҢЩҶ ШЁЩ„ЩҲЪ© ШұЩҲ ЩҫШ§Ъ© Ъ©ЩҶ                               в•‘
# в•‘  + Ш®Ш· translate ШұЩҲ Ш§ШІ main_menu ЩҫШ§Ъ© Ъ©ЩҶ                      в•‘
# в•‘  + Ш®Ш· handle_translate ШұЩҲ Ш§ШІ ROUTER ЩҫШ§Ъ© Ъ©ЩҶ                  в•‘
# в•‘  + Ш®Ш· state translate ШұЩҲ Ш§ШІ message_handler ЩҫШ§Ъ© Ъ©ЩҶ          в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ

def translate_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹ”Ө ШӘШұШ¬Щ…ЩҮ Щ…ШӘЩҶ", callback_data="tr_start")],
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_main")],
    ])

async def do_translate(text):
    try:
        async with httpx.AsyncClient(
            timeout=20
        ) as client:
            persian_chars = sum(1 for c in text if '\u0600' <= c <= '\u06FF')
            is_persian = persian_chars > len(text) * 0.3

            if is_persian:
                target_lang = "en"
                direction = "рҹҮ®рҹҮ· ЩҒШ§ШұШіЫҢ вҶ’ рҹҮ¬рҹҮ§ English"
            else:
                target_lang = "fa"
                direction = "рҹҮ¬рҹҮ§ English вҶ’ рҹҮ®рҹҮ· ЩҒШ§ШұШіЫҢ"

            response = await client.get(
                "https://api.mymemory.translated.net/get",
                params={
                    "q": text,
                    "langpair": f"{'fa' if is_persian else 'en'}|{target_lang}"
                }
            )
            data = response.json()
            translated = data["responseData"]["translatedText"]

            # ШұЩҒШ№ Ъ©Ш§ШұШ§Ъ©ШӘШұЩҮШ§ЫҢ HTML Щ…Ш«Щ„ &#10; (Ш®Ш· Ш¬ШҜЫҢШҜ) ЩҲ &amp; ЩҲ...
            import html
            translated = html.unescape(translated)

            return direction, translated

    except Exception as e:
        print(f"Translate error: {e}")
        return None, None

async def handle_translate(query, user_id):
    data = query.data

    if data == "menu_translate":
        await query.edit_message_text(
            "рҹҢҗ Щ…Ш§ЪҳЩҲЩ„ ШӘШұШ¬Щ…ЩҮ\n\nЩ…ШӘЩҶ Ш®ЩҲШҜ ШұШ§ ШЁЩҶЩҲЫҢШіЫҢШҜ вҖ” Ш§ЪҜШұ ЩҒШ§ШұШіЫҢ ШЁШ§ШҙШҜ ШЁЩҮ Ш§ЩҶЪҜЩ„ЫҢШіЫҢ ЩҲ Ш§ЪҜШұ Ш§ЩҶЪҜЩ„ЫҢШіЫҢ ШЁШ§ШҙШҜ ШЁЩҮ ЩҒШ§ШұШіЫҢ ШӘШұШ¬Щ…ЩҮ Щ…ЫҢвҖҢШҙЩҲШҜ.",
            reply_markup=translate_menu()
        )

    elif data == "tr_start":
        set_state(user_id, "tr_waiting")
        await query.edit_message_text(
            "вңҚпёҸ Щ…ШӘЩҶ Щ…ЩҲШұШҜ ЩҶШёШұ ШұШ§ ШӘШ§ЫҢЩҫ Ъ©ЩҶЫҢШҜ:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="menu_translate")]
            ])
        )

async def handle_translate_message(user_id, text, update):
    state, _ = get_state(user_id)

    if state == "tr_waiting":
        await update.message.reply_text("вҸі ШҜШұ ШӯШ§Щ„ ШӘШұШ¬Щ…ЩҮ...")
        direction, translated = await do_translate(text)

        if translated:
            await update.message.reply_text(
                f"рҹҢҗ {direction}\n\n"
                f"рҹ“қ Щ…ШӘЩҶ Ш§ШөЩ„ЫҢ:\n{text}\n\n"
                f"вң… ШӘШұШ¬Щ…ЩҮ:\n{translated}",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("рҹ”„ ШӘШұШ¬Щ…ЩҮ Ш¬ШҜЫҢШҜ", callback_data="tr_start")],
                    [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ ШЁЩҮ Щ…ЩҶЩҲ", callback_data="menu_translate")],
                ])
            )
        else:
            await update.message.reply_text(
                "вқҢ Ш®Ш·Ш§ ШҜШұ ШӘШұШ¬Щ…ЩҮ. ШҜЩҲШЁШ§ШұЩҮ Ш§Щ…ШӘШӯШ§ЩҶ Ъ©ЩҶЫҢШҜ.",
                reply_markup=translate_menu()
            )
        clear_state(user_id)
        return True

    return False

# в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ ЩҫШ§ЫҢШ§ЩҶ MODULE: TRANSLATE в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ


# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                  MODULE: POSITION SIZE                       в•‘
# в•‘  ШЁШұШ§ЫҢ ШӯШ°ЩҒ: Ш§ЫҢЩҶ ШЁЩ„ЩҲЪ© ШұЩҲ ЩҫШ§Ъ© Ъ©ЩҶ                               в•‘
# в•‘  + Ш®Ш· ps_size ШұЩҲ Ш§ШІ financial_menu ЩҫШ§Ъ© Ъ©ЩҶ                   в•‘
# в•‘  + Ш®Ш· ps_ ШұЩҲ Ш§ШІ ROUTER Ш§ШөЩ„ЫҢ ЩҫШ§Ъ© Ъ©ЩҶ                          в•‘
# в•‘  + state ЩҮШ§ЫҢ ps_ ШұЩҲ Ш§ШІ message_handler ЩҫШ§Ъ© Ъ©ЩҶ               в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ

# в”Җв”Җв”Җ Щ…ШҙШ®ШөШ§ШӘ ЩӮШұШ§ШұШҜШ§ШҜЫҢ ЩҮШұ ШЁШ§ШІШ§Шұ ШЁШұШ§ЫҢ Щ…ШӯШ§ШіШЁЩҮ Ш§ШұШІШҙ ЩҮШұ ЩҲШ§ШӯШҜ ШӯШұЪ©ШӘ ЩӮЫҢЩ…ШӘ в”Җв”Җ
# pip_size: Ъ©ЩҲЪҶЪ©вҖҢШӘШұЫҢЩҶ ЩҲШ§ШӯШҜ ШӯШұЪ©ШӘ ЩӮЫҢЩ…ШӘ Ш§ШіШӘШ§ЩҶШҜШ§ШұШҜ ШўЩҶ ЩҶЩ…Ш§ШҜ
# contract_size: Ш§ЩҶШҜШ§ШІЩҮ Ыұ Щ„Ш§ШӘ Ш§ШіШӘШ§ЩҶШҜШ§ШұШҜ (ШӘШ№ШҜШ§ШҜ ЩҲШ§ШӯШҜ Ш§ШұШІ ЩҫШ§ЫҢЩҮ/Ъ©Ш§Щ„Ш§)
FX_CONTRACT_SPECS = {
    "EURUSD": {"pip_size": 0.0001, "contract_size": 100000, "quote_is_usd": True},
    "GBPUSD": {"pip_size": 0.0001, "contract_size": 100000, "quote_is_usd": True},
    "AUDUSD": {"pip_size": 0.0001, "contract_size": 100000, "quote_is_usd": True},
    "NZDUSD": {"pip_size": 0.0001, "contract_size": 100000, "quote_is_usd": True},
    "USDCAD": {"pip_size": 0.0001, "contract_size": 100000, "quote_is_usd": False},
    "USDJPY": {"pip_size": 0.01,   "contract_size": 100000, "quote_is_usd": False},
    "USDCHF": {"pip_size": 0.0001, "contract_size": 100000, "quote_is_usd": False},
    "EURGBP": {"pip_size": 0.0001, "contract_size": 100000, "quote_is_usd": False},
    "EURJPY": {"pip_size": 0.01,   "contract_size": 100000, "quote_is_usd": False},
    "GBPJPY": {"pip_size": 0.01,   "contract_size": 100000, "quote_is_usd": False},
}
METAL_CONTRACT_SPECS = {
    "GOLD": {"contract_size": 100},     # 1 Щ„Ш§ШӘ Ш·Щ„Ш§ = 100 Ш§ЩҲЩҶШі
    "XAUUSD": {"contract_size": 100},
    "SILVER": {"contract_size": 5000},  # 1 Щ„Ш§ШӘ ЩҶЩӮШұЩҮ = 5000 Ш§ЩҲЩҶШі
    "XAGUSD": {"contract_size": 5000},
}

def ps_classify_symbol(symbol):
    """ШӘШҙШ®ЫҢШө ЩҶЩҲШ№ ЩҶЩ…Ш§ШҜ ШЁШұШ§ЫҢ Ш§ЩҶШӘШ®Ш§ШЁ ЩҒШұЩ…ЩҲЩ„ Щ…ШӯШ§ШіШЁЩҮ: forex_pair / metal / crypto / iran / forex_other"""
    s = symbol.upper()
    if s in FX_CONTRACT_SPECS:
        return "forex_pair"
    if s in METAL_CONTRACT_SPECS:
        return "metal"
    if s in CRYPTO_SYMBOLS:
        return "crypto"
    if s in IRAN_SYMBOLS:
        return "iran"
    return "forex_other"

async def ps_calculate_position(symbol, direction, entry_price, sl_price, risk_usd):
    """
    Щ…ШӯШ§ШіШЁЩҮ ШӯШ¬Щ… ЩҫЫҢШҙЩҶЩҮШ§ШҜЫҢ ШЁШұ Ш§ШіШ§Ші Щ…ЩӮШҜШ§Шұ ШұЫҢШіЪ© ШҜЩ„Ш§ШұЫҢ.
    Ш®ШұЩҲШ¬ЫҢ: dict ШҙШ§Щ…Щ„ risk_distance, unit_value, lots/qty, explanation
    """
    symbol = symbol.upper()
    category = ps_classify_symbol(symbol)
    risk_distance = abs(entry_price - sl_price)

    if risk_distance == 0:
        return None

    if category == "forex_pair":
        spec = FX_CONTRACT_SPECS[symbol]
        contract_size = spec["contract_size"]
        if spec["quote_is_usd"]:
            # Щ…Ш«Щ„ EURUSD: Ш§ШұШІШҙ ШӯШұЪ©ШӘ ЩӮЫҢЩ…ШӘ ШЁШұШ§ЫҢ 1 Щ„Ш§ШӘ Ъ©Ш§Щ…Щ„ = ЩҒШ§ШөЩ„ЩҮ ЩӮЫҢЩ…ШӘ Г— Ш§ЩҶШҜШ§ШІЩҮ ЩӮШұШ§ШұШҜШ§ШҜ
            value_per_lot = risk_distance * contract_size
        else:
            # Щ…Ш«Щ„ USDJPY/USDCAD: Ш§ШұШІ Quote ШҜЩ„Ш§Шұ ЩҶЫҢШіШӘШҢ ШЁШ§ЫҢШҜ ШӘЩӮШіЫҢЩ… ШЁШұ ЩҶШұШ® Щ„ШӯШёЩҮвҖҢШ§ЫҢ Ш¬ЩҒШӘвҖҢШ§ШұШІ ШҙЩҲШҜ
            # ШӘШ§ ШӯШұЪ©ШӘ ЩӮЫҢЩ…ШӘ (Ъ©ЩҮ ШЁЩҮ ЩҲШ§ШӯШҜ Quote Ш§ШіШӘ) ШЁЩҮ ШҜЩ„Ш§Шұ ШӘШЁШҜЫҢЩ„ ШҙЩҲШҜ
            live_price, _ = await get_price_by_symbol(symbol, "forex")
            reference_price = live_price if live_price else entry_price
            value_per_lot = (risk_distance * contract_size) / reference_price
        lots = risk_usd / value_per_lot
        return {
            "category": "forex_pair", "risk_distance": risk_distance,
            "lots": lots, "value_per_lot": value_per_lot, "unit": "Щ„Ш§ШӘ Ш§ШіШӘШ§ЩҶШҜШ§ШұШҜ"
        }

    elif category == "metal":
        spec = METAL_CONTRACT_SPECS[symbol]
        contract_size = spec["contract_size"]
        value_per_lot = risk_distance * contract_size
        lots = risk_usd / value_per_lot
        return {
            "category": "metal", "risk_distance": risk_distance,
            "lots": lots, "value_per_lot": value_per_lot, "unit": "Щ„Ш§ШӘ Ш§ШіШӘШ§ЩҶШҜШ§ШұШҜ"
        }

    elif category == "crypto":
        qty = risk_usd / risk_distance
        return {
            "category": "crypto", "risk_distance": risk_distance,
            "lots": qty, "value_per_lot": risk_distance, "unit": f"ЩҲШ§ШӯШҜ {symbol}"
        }

    elif category == "iran":
        qty = risk_usd / risk_distance
        return {
            "category": "iran", "risk_distance": risk_distance,
            "lots": qty, "value_per_lot": risk_distance, "unit": "ЩҲШ§ШӯШҜ (ШұЫҢШ§Щ„вҖҢЩ…ШӯЩҲШұ)"
        }

    else:  # forex_other - Ш¬ЩҒШӘ Ш§ШұШІ ЩҶШ§ШҙЩҶШ§Ш®ШӘЩҮ ШЁШ§ pip Ш§ШіШӘШ§ЩҶШҜШ§ШұШҜ ЩҒШұШ¶ Щ…ЫҢвҖҢШҙЩҲШҜ
        contract_size = 100000
        live_price, _ = await get_price_by_symbol(symbol, "forex")
        reference_price = live_price if live_price else entry_price
        value_per_lot = (risk_distance * contract_size) / reference_price if reference_price else 0
        lots = risk_usd / value_per_lot if value_per_lot else 0
        return {
            "category": "forex_other", "risk_distance": risk_distance,
            "lots": lots, "value_per_lot": value_per_lot, "unit": "Щ„Ш§ШӘ Ш§ШіШӘШ§ЩҶШҜШ§ШұШҜ (ШӘЩӮШұЫҢШЁЫҢ)"
        }

# в”Җв”Җв”Җ Щ…ЩҶЩҲЩҮШ§ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
def position_size_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹ“җ Щ…ШӯШ§ШіШЁЩҮ Ш¬ШҜЫҢШҜ", callback_data="ps_new")],
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_financial")],
    ])

def ps_direction_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹ“Ҳ Long (Ш®ШұЫҢШҜ)", callback_data="ps_dir_long")],
        [InlineKeyboardButton("рҹ“ү Short (ЩҒШұЩҲШҙ)", callback_data="ps_dir_short")],
    ])

def ps_risk_type_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹ’ө Щ…ЩӮШҜШ§Шұ ШҜЩ„Ш§ШұЫҢ Ш«Ш§ШЁШӘ", callback_data="ps_risk_fixed")],
        [InlineKeyboardButton("рҹ“Ҡ ШҜШұШөШҜЫҢ Ш§ШІ ШЁШ§Щ„Ш§ЩҶШі", callback_data="ps_risk_percent")],
    ])

# в”Җв”Җв”Җ ЩҮЩҶШҜЩ„Шұ ШҜЪ©Щ…ЩҮвҖҢЩҮШ§ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
async def handle_position_size(query, user_id):
    data = query.data

    if data == "ps_size":
        await query.edit_message_text(
            "рҹ“җ Щ…ШӯШ§ШіШЁЩҮ Щ…ЫҢШІШ§ЩҶ Ш®ШұЫҢШҜ (Position Sizing)\n\n"
            "ШЁШ§ ЪҜШұЩҒШӘЩҶ ЩҶЩӮШ·ЩҮ ЩҲШұЩҲШҜШҢ ШӯШҜ Ш¶ШұШұ ЩҲ Щ…ЫҢШІШ§ЩҶ ШұЫҢШіЪ©ЫҢ Ъ©ЩҮ Щ…ЫҢвҖҢШ®ЩҲШ§ЩҮЫҢШҜ "
            "ШЁЩҫШ°ЫҢШұЫҢШҜ (ШҜЩ„Ш§ШұЫҢ ЫҢШ§ ШҜШұШөШҜЫҢ Ш§ШІ ШЁШ§Щ„Ш§ЩҶШі)ШҢ ШӯШ¬Щ… Щ…ЩҶШ§ШіШЁ Щ…Ш№Ш§Щ…Щ„ЩҮ "
            "(Щ„Ш§ШӘ Ш§ШіШӘШ§ЩҶШҜШ§ШұШҜ ЫҢШ§ Щ…ЩӮШҜШ§Шұ ЩҲШ§ШӯШҜ) ШұШ§ Щ…ШӯШ§ШіШЁЩҮ Щ…ЫҢвҖҢЪ©ЩҶШҜ.",
            reply_markup=position_size_menu()
        )

    elif data == "ps_new":
        set_state(user_id, "ps_symbol")
        await query.edit_message_text(
            "рҹ“җ Щ…ШӯШ§ШіШЁЩҮ Ш¬ШҜЫҢШҜ\n\nЩҶЩ…Ш§ШҜ Ш§ШұШІ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:\nЩ…Ш«Ш§Щ„: `EURUSD` `GOLD` `BTC` `USD`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="ps_size")]])
        )

    elif data in ["ps_dir_long", "ps_dir_short"]:
        direction = "long" if data == "ps_dir_long" else "short"
        state, sdata = get_state(user_id)
        if state == "ps_direction":
            set_state(user_id, "ps_entry_price", f"{sdata}|{direction}")
            await query.edit_message_text(
                f"вң… Ш¬ЩҮШӘ: {'Long рҹ“Ҳ' if direction=='long' else 'Short рҹ“ү'}\n\nЩӮЫҢЩ…ШӘ ЩҲШұЩҲШҜ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:"
            )

    elif data in ["ps_risk_fixed", "ps_risk_percent"]:
        state, sdata = get_state(user_id)
        if state == "ps_risk_type":
            if data == "ps_risk_fixed":
                set_state(user_id, "ps_risk_amount", f"{sdata}|fixed")
                await query.edit_message_text("рҹ’ө Щ…ЩӮШҜШ§Шұ ШұЫҢШіЪ© ШЁЩҮ ШҜЩ„Ш§Шұ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:\nЩ…Ш«Ш§Щ„: `10`", parse_mode="Markdown")
            else:
                set_state(user_id, "ps_balance", f"{sdata}|percent")
                await query.edit_message_text("рҹ’° ШЁШ§Щ„Ш§ЩҶШі ШӯШіШ§ШЁ ШұШ§ ШЁЩҮ ШҜЩ„Ш§Шұ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:\nЩ…Ш«Ш§Щ„: `1000`", parse_mode="Markdown")

async def handle_position_size_message(user_id, text, update):
    state, data = get_state(user_id)
    text = text.strip()

    if state == "ps_symbol":
        symbol = text.upper()
        set_state(user_id, "ps_direction", symbol)
        await update.message.reply_text("Ш¬ЩҮШӘ Щ…Ш№Ш§Щ…Щ„ЩҮ ШұШ§ Ш§ЩҶШӘШ®Ш§ШЁ Ъ©ЩҶЫҢШҜ:", reply_markup=ps_direction_menu())
        return True

    elif state == "ps_entry_price":
        try:
            entry_price = float(text.replace(",", ""))
            set_state(user_id, "ps_sl_price", f"{data}|{entry_price}")
            await update.message.reply_text("ЩӮЫҢЩ…ШӘ ШӯШҜ Ш¶ШұШұ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        except:
            await update.message.reply_text("вқҢ Щ„Ш·ЩҒШ§ЩӢ ЫҢЪ© Ш№ШҜШҜ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        return True

    elif state == "ps_sl_price":
        try:
            sl_price = float(text.replace(",", ""))
            parts = data.split("|")
            symbol, direction, entry_price = parts[0], parts[1], float(parts[2])

            if (direction == "long" and sl_price >= entry_price) or (direction == "short" and sl_price <= entry_price):
                await update.message.reply_text("вқҢ ШӯШҜ Ш¶ШұШұ ШЁШ§ Ш¬ЩҮШӘ Щ…Ш№Ш§Щ…Щ„ЩҮ ЩҮЩ…Ш®ЩҲШ§ЩҶЫҢ ЩҶШҜШ§ШұШҜ.\nШҜЩҲШЁШ§ШұЩҮ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
                return True

            set_state(user_id, "ps_risk_type", f"{symbol}|{direction}|{entry_price}|{sl_price}")
            await update.message.reply_text("ШұЩҲШҙ ШӘШ№ЫҢЫҢЩҶ ШұЫҢШіЪ© ШұШ§ Ш§ЩҶШӘШ®Ш§ШЁ Ъ©ЩҶЫҢШҜ:", reply_markup=ps_risk_type_menu())
        except:
            await update.message.reply_text("вқҢ Щ„Ш·ЩҒШ§ЩӢ ЫҢЪ© Ш№ШҜШҜ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        return True

    elif state == "ps_risk_amount":
        try:
            risk_usd = float(text.replace(",", ""))
            if risk_usd <= 0:
                await update.message.reply_text("вқҢ Щ…ЩӮШҜШ§Шұ ШұЫҢШіЪ© ШЁШ§ЫҢШҜ ШЁШІШұЪҜвҖҢШӘШұ Ш§ШІ ШөЩҒШұ ШЁШ§ШҙШҜ:")
                return True
            parts = data.split("|")
            symbol, direction, entry_price, sl_price, risk_mode = parts
            await run_position_size_final(user_id, update, symbol, direction,
                                           float(entry_price), float(sl_price), risk_usd)
        except:
            await update.message.reply_text("вқҢ Щ„Ш·ЩҒШ§ЩӢ ЫҢЪ© Ш№ШҜШҜ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        return True

    elif state == "ps_balance":
        try:
            balance = float(text.replace(",", ""))
            if balance <= 0:
                await update.message.reply_text("вқҢ ШЁШ§Щ„Ш§ЩҶШі ШЁШ§ЫҢШҜ ШЁШІШұЪҜвҖҢШӘШұ Ш§ШІ ШөЩҒШұ ШЁШ§ШҙШҜ:")
                return True
            set_state(user_id, "ps_percent", f"{data}|{balance}")
            await update.message.reply_text("ШҜШұШөШҜ ШұЫҢШіЪ© Щ…Ш¬Ш§ШІ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:\nЩ…Ш«Ш§Щ„: `2`")
        except:
            await update.message.reply_text("вқҢ Щ„Ш·ЩҒШ§ЩӢ ЫҢЪ© Ш№ШҜШҜ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        return True

    elif state == "ps_percent":
        try:
            percent = float(text.replace(",", ""))
            if percent <= 0 or percent > 100:
                await update.message.reply_text("вқҢ ШҜШұШөШҜ ШЁШ§ЫҢШҜ ШЁЫҢЩҶ 0 ЩҲ 100 ШЁШ§ШҙШҜ:")
                return True
            parts = data.split("|")
            symbol, direction, entry_price, sl_price, risk_mode, balance = parts
            risk_usd = float(balance) * (percent / 100)
            await run_position_size_final(user_id, update, symbol, direction,
                                           float(entry_price), float(sl_price), risk_usd,
                                           balance=float(balance), percent=percent)
        except:
            await update.message.reply_text("вқҢ Щ„Ш·ЩҒШ§ЩӢ ЫҢЪ© Ш№ШҜШҜ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        return True

    return False

async def run_position_size_final(user_id, update, symbol, direction, entry_price, sl_price,
                                    risk_usd, balance=None, percent=None):
    await update.message.reply_text("вҸі ШҜШұ ШӯШ§Щ„ Щ…ШӯШ§ШіШЁЩҮ...")

    result = await ps_calculate_position(symbol, direction, entry_price, sl_price, risk_usd)
    if not result:
        await update.message.reply_text("вқҢ Ш®Ш·Ш§ ШҜШұ Щ…ШӯШ§ШіШЁЩҮ (ЩҒШ§ШөЩ„ЩҮ ШұЫҢШіЪ© ШөЩҒШұ Ш§ШіШӘ).", reply_markup=position_size_menu())
        clear_state(user_id)
        return

    risk_line = f"рҹ’ө Щ…ЩӮШҜШ§Шұ ШұЫҢШіЪ©: ${risk_usd:,.2f}"
    if balance and percent:
        risk_line += f"\nрҹ’° ШЁШ§Щ„Ш§ЩҶШі: ${balance:,.2f} | ШҜШұШөШҜ ШұЫҢШіЪ©: {percent}%"

    text = (
        f"рҹ“җ Щ…ШӯШ§ШіШЁЩҮ Щ…ЫҢШІШ§ЩҶ Ш®ШұЫҢШҜ вҖ” {symbol} ({'Long' if direction=='long' else 'Short'})\n"
        f"`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`\n"
        f"ЩҲШұЩҲШҜ: {entry_price}\n"
        f"ШӯШҜ Ш¶ШұШұ: {sl_price}\n"
        f"ЩҒШ§ШөЩ„ЩҮ ШұЫҢШіЪ©: {result['risk_distance']:.5f}\n"
        f"`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`\n"
        f"{risk_line}\n"
        f"`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`\n"
        f"рҹ“Ұ ШӯШ¬Щ… ЩҫЫҢШҙЩҶЩҮШ§ШҜЫҢ: {result['lots']:.4f} {result['unit']}\n"
        f"`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`"
    )

    if result["category"] == "forex_other":
        text += "\n\nвҡ пёҸ Ш§ЫҢЩҶ ЩҶЩ…Ш§ШҜ ШҜШұ Щ„ЫҢШіШӘ ШҜЩӮЫҢЩӮ ШұШЁШ§ШӘ ЩҶШЁЩҲШҜШӣ Щ…ШӯШ§ШіШЁЩҮ ШЁШ§ ЩҒШұШ¶ pip Ш§ШіШӘШ§ЩҶШҜШ§ШұШҜ ЩҒШ§ШұЪ©Ші Ш§ЩҶШ¬Ш§Щ… ШҙШҜЩҮ ЩҲ Щ…Щ…Ъ©ЩҶ Ш§ШіШӘ ШӘЩӮШұЫҢШЁЫҢ ШЁШ§ШҙШҜ."

    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=position_size_menu())
    clear_state(user_id)

# в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ ЩҫШ§ЫҢШ§ЩҶ MODULE: POSITION SIZE в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ


# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                  MODULE: BACKTEST                            в•‘
# в•‘  ШЁШұШ§ЫҢ ШӯШ°ЩҒ: Ш§ЫҢЩҶ ШЁЩ„ЩҲЪ© ШұЩҲ ЩҫШ§Ъ© Ъ©ЩҶ                               в•‘
# в•‘  + Ш®Ш· backtest ШұЩҲ Ш§ШІ financial_menu ЩҫШ§Ъ© Ъ©ЩҶ                  в•‘
# в•‘  + Ш®Ш· fin_backtest/bt_ ШұЩҲ Ш§ШІ ROUTER Ш§ШөЩ„ЫҢ ЩҫШ§Ъ© Ъ©ЩҶ             в•‘
# в•‘  + state ЩҮШ§ЫҢ bt_ ШұЩҲ Ш§ШІ message_handler ЩҫШ§Ъ© Ъ©ЩҶ               в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

IRAN_TZ_OFFSET = timedelta(hours=3, minutes=30)
R_LEVELS = [0.5, 1, 1.5, 2, 3, 4, 8, 10]

def describe_strategy(rules):
    """ШӘЩҲЩ„ЫҢШҜ Ш®Щ„Ш§ШөЩҮ Щ…ШӘЩҶЫҢ Ш®ЩҲШ§ЩҶШ§ Ш§ШІ ЩӮЩҲШ§ЩҶЫҢЩҶ Ш§ШіШӘШұШ§ШӘЪҳЫҢ ШЁШұШ§ЫҢ ЩҶЩ…Ш§ЫҢШҙ ШЁЩҮ Ъ©Ш§ШұШЁШұ"""
    lines = []
    for r in R_LEVELS[:-1]:  # R10 ЩҮЩ…ЫҢШҙЩҮ Ш®ШұЩҲШ¬ Ъ©Ш§Щ…Щ„ Ш§ШіШӘШҢ Ш¬ШІЩҲ ШіЩҲШ§Щ„Ш§ШӘ ЩҶЫҢШіШӘ
        key = str(r)
        rule = rules.get(key)
        if isinstance(rule, dict) and rule.get("type") == "r_trail":
            value = rule["value"]
            if value == 0:
                lines.append(f"рҹ”ё R{r} вҶ’ ШұЫҢШіЪ©вҖҢЩҒШұЫҢ (SL = ЩҶЩӮШ·ЩҮ ЩҲШұЩҲШҜ)")
            else:
                lines.append(f"рҹ”ё R{r} вҶ’ SL ШЁЩҮ R{value} Щ…ЩҶШӘЩӮЩ„ Щ…ЫҢвҖҢШҙЩҲШҜ")
        else:
            lines.append(f"рҹ”ё R{r} вҶ’ ШЁШҜЩҲЩҶ ШӘШәЫҢЫҢШұ (ШұШҜ Щ…ЫҢвҖҢШҙЩҲШҜ)")
    lines.append("рҹ”ё R10 вҶ’ Ш®ШұЩҲШ¬ Ъ©Ш§Щ…Щ„ (ЩҮЩ…ЫҢШҙЩҮ Ш«Ш§ШЁШӘ)")
    return "\n".join(lines)

def to_iran_time(dt_naive_utc_like):
    """ШӘШЁШҜЫҢЩ„ datetime (Ъ©ЩҮ Ш§ШІ Yahoo Щ…ЫҢвҖҢШўЫҢШҜШҢ Щ…Ш№Щ…ЩҲЩ„Ш§ЩӢ UTC) ШЁЩҮ ШіШ§Ш№ШӘ Ш§ЫҢШұШ§ЩҶ ШЁШұШ§ЫҢ ЩҶЩ…Ш§ЫҢШҙ"""
    return dt_naive_utc_like + IRAN_TZ_OFFSET

# в”Җв”Җв”Җ ШҜЫҢШӘШ§ШЁЫҢШі Ш§ШіШӘШұШ§ШӘЪҳЫҢ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
def init_backtest_db():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS bt_strategies (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id    INTEGER,
        name       TEXT,
        rules_json TEXT,
        created_at TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS bt_results (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id     INTEGER,
        batch_id    TEXT,
        symbol      TEXT,
        direction   TEXT,
        entry_time  TEXT,
        entry_price REAL,
        sl_price    REAL,
        strategy    TEXT,
        max_r       REAL,
        result_r    REAL,
        exit_time   TEXT,
        exit_price  REAL,
        status      TEXT,
        created_at  TEXT
    )''')
    conn.commit()
    conn.close()

def save_strategy(user_id, name, rules):
    import json
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO bt_strategies (user_id, name, rules_json, created_at) VALUES (?, ?, ?, ?)",
              (user_id, name, json.dumps(rules), datetime.now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def get_strategies(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id, name, rules_json FROM bt_strategies WHERE user_id=? ORDER BY id DESC", (user_id,))
    rows = c.fetchall()
    conn.close()
    return rows

def get_strategy_by_id(sid):
    import json
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT name, rules_json FROM bt_strategies WHERE id=?", (sid,))
    row = c.fetchone()
    conn.close()
    if row:
        return row[0], json.loads(row[1])
    return None, None

def save_bt_result(user_id, batch_id, symbol, direction, entry_time, entry_price,
                    sl_price, strategy_name, max_r, result_r, exit_time, exit_price, status):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''INSERT INTO bt_results
        (user_id, batch_id, symbol, direction, entry_time, entry_price, sl_price,
         strategy, max_r, result_r, exit_time, exit_price, status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (user_id, batch_id, symbol, direction, entry_time, entry_price, sl_price,
         strategy_name, max_r, result_r, exit_time, exit_price, status,
         datetime.now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def get_batch_results(user_id, batch_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT * FROM bt_results WHERE user_id=? AND batch_id=? ORDER BY id", (user_id, batch_id))
    rows = c.fetchall()
    conn.close()
    return rows

# в”Җв”Җв”Җ ШҜШұЫҢШ§ЩҒШӘ Ъ©ЩҶШҜЩ„вҖҢЩҮШ§ЫҢ ШӘШ§ШұЫҢШ®ЫҢ (ШЁШҜЩҲЩҶ ШӘШ§ЫҢЩ…вҖҢЩҒШұЫҢЩ… ШұЩҲШІШ§ЩҶЩҮ) в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
async def get_historical_candles(symbol, entry_date, entry_hm, interval):
    """entry_date: YYYY-MM-DD | entry_hm: HH:MM (ШЁЩҮ ЩҲЩӮШӘ Ш§ЫҢШұШ§ЩҶ) | interval: 1m/15m/1h"""
    try:
        entry_dt_iran = datetime.strptime(f"{entry_date} {entry_hm}", "%Y-%m-%d %H:%M")
        entry_dt_utc = entry_dt_iran - IRAN_TZ_OFFSET
        days_ago = (datetime.now() - entry_dt_utc).days

        limits = {"1m": 7, "15m": 60, "1h": 730}
        max_days = limits.get(interval)
        if max_days is None or days_ago > max_days:
            return None, None, f"вҡ пёҸ ШӘШ§ЫҢЩ…вҖҢЩҒШұЫҢЩ… Ш§ЩҶШӘШ®Ш§ШЁЫҢ ЩҒЩӮШ· ШӘШ§ {max_days} ШұЩҲШІ ЪҜШ°ШҙШӘЩҮ ШұШ§ ЩҫШҙШӘЫҢШЁШ§ЩҶЫҢ Щ…ЫҢвҖҢЪ©ЩҶШҜ.\nШӘШ§ШұЫҢШ® ШҙЩ…Ш§ {days_ago} ШұЩҲШІ ЩҫЫҢШҙ Ш§ШіШӘ."

        range_map = {
            "1m": f"{min(days_ago + 1, 7)}d",
            "15m": f"{min(days_ago + 1, 60)}d",
            "1h": f"{min(days_ago + 5, 730)}d",
        }
        range_param = range_map[interval]

        yahoo_sym = FOREX_YAHOO.get(symbol.upper())
        if not yahoo_sym and symbol.upper() in GOLDAPI_SYMBOLS:
            yahoo_sym = "GC%3DF" if symbol.upper() in ["GOLD", "XAUUSD"] else "SI%3DF"
        if not yahoo_sym:
            yahoo_sym = symbol.upper() + "%3DX"

        async with httpx.AsyncClient(timeout=15, headers={"User-Agent": "Mozilla/5.0"}) as client:
            r = await client.get(
                f"https://query1.finance.yahoo.com/v8/finance/chart/{yahoo_sym}",
                params={"interval": interval, "range": range_param}
            )
            data = r.json()["chart"]["result"][0]
            timestamps = data["timestamp"]
            quotes = data["indicators"]["quote"][0]

            candles = []
            entry_ts = entry_dt_utc.timestamp()
            for i, ts in enumerate(timestamps):
                if ts >= entry_ts:
                    high = quotes["high"][i]
                    low = quotes["low"][i]
                    close = quotes["close"][i]
                    if high is not None and low is not None:
                        candles.append({
                            "time_utc": datetime.fromtimestamp(ts),
                            "high": high, "low": low, "close": close
                        })
            if not candles:
                return None, None, "вқҢ ШҜШ§ШҜЩҮвҖҢШ§ЫҢ ШЁШұШ§ЫҢ Ш§ЫҢЩҶ ШЁШ§ШІЩҮ ШІЩ…Ш§ЩҶЫҢ ЫҢШ§ЩҒШӘ ЩҶШҙШҜ."
            return candles, entry_dt_utc, None
    except Exception as e:
        print(f"Historical data error: {e}")
        return None, None, "вқҢ Ш®Ш·Ш§ ШҜШұ ШҜШұЫҢШ§ЩҒШӘ ШҜШ§ШҜЩҮвҖҢЩҮШ§ЫҢ ШӘШ§ШұЫҢШ®ЫҢ."

# в”Җв”Җв”Җ Щ…ЩҲШӘЩҲШұ ШҙШЁЫҢЩҮвҖҢШіШ§ШІЫҢ ШЁШ§ ЩӮЩҲШ§ЩҶЫҢЩҶ ШіЩҒШ§ШұШҙЫҢ Ъ©Ш§ШұШЁШұ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
def run_custom_simulation(candles, entry_price, sl_price, direction, strategy_rules):
    """
    strategy_rules: dict Щ…Ш«Щ„ {"0.5": None, "1": {"type": "r_trail", "value": 0.2}, ...}
    Щ…ЩӮШҜШ§Шұ None ЫҢШ№ЩҶЫҢ SL ШӘШәЫҢЫҢШұ ЩҶЩ…ЫҢвҖҢЪ©ЩҶШҜ ШҜШұ ШўЩҶ ШіШ·Шӯ
    Щ…ЩӮШҜШ§Шұ {"type": "r_trail", "value": X} ЫҢШ№ЩҶЫҢ SL ШұЩҲЫҢ ШіШ·Шӯ RX ЩӮШұШ§Шұ Щ…ЫҢвҖҢЪҜЫҢШұШҜ
        (Щ…Ш«Щ„Ш§ЩӢ value=0 ЫҢШ№ЩҶЫҢ ЩҶЩӮШ·ЩҮ ЩҲШұЩҲШҜШҢ value=0.2 ЫҢШ№ЩҶЫҢ R0.2 Ш§ШІ ЩҶЩӮШ·ЩҮ ЩҲШұЩҲШҜ)
    ШҜШұ R10 ЩҮЩ…ЫҢШҙЩҮ Ш®ШұЩҲШ¬ Ъ©Ш§Щ…Щ„
    """
    R = abs(entry_price - sl_price)
    is_long = direction == "long"

    def level(r_mult):
        return entry_price + (r_mult * R) if is_long else entry_price - (r_mult * R)

    current_sl = sl_price
    max_r_reached = 0
    triggered_levels = set()

    for candle in candles:
        high, low = candle["high"], candle["low"]

        hit_sl = (low <= current_sl) if is_long else (high >= current_sl)
        if hit_sl:
            exit_r = (current_sl - entry_price) / R if is_long else (entry_price - current_sl) / R
            return {
                "exit_price": current_sl, "exit_time_utc": candle["time_utc"],
                "result_r": round(exit_r, 2), "max_r": max_r_reached, "status": "closed"
            }

        reached = high if is_long else low
        for r_mult in R_LEVELS:
            if r_mult in triggered_levels:
                continue
            target_level = level(r_mult)
            condition = (reached >= target_level) if is_long else (reached <= target_level)
            if condition:
                triggered_levels.add(r_mult)
                max_r_reached = r_mult

                if r_mult == 10:
                    return {
                        "exit_price": level(10), "exit_time_utc": candle["time_utc"],
                        "result_r": 10.0, "max_r": 10.0, "status": "closed"
                    }

                rule = strategy_rules.get(str(r_mult))
                if isinstance(rule, dict) and rule.get("type") == "r_trail":
                    current_sl = level(rule["value"])
                # rule None -> ШЁШҜЩҲЩҶ ШӘШәЫҢЫҢШұ

    last_close = candles[-1]["close"] if candles else entry_price
    open_r = (last_close - entry_price) / R if is_long else (entry_price - last_close) / R
    return {
        "exit_price": None, "exit_time_utc": None,
        "result_r": round(open_r, 2), "max_r": max_r_reached, "status": "open"
    }

# в”Җв”Җв”Җ ШіШ§Ш®ШӘ ЩҒШ§ЫҢЩ„ Ш§Ъ©ШіЩ„ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
def build_excel_report(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Backtest Results"

    headers = ["ЩҶЩ…Ш§ШҜ", "Ш¬ЩҮШӘ", "ШӘШ§ШұЫҢШ® ЩҲШұЩҲШҜ (Ш§ЫҢШұШ§ЩҶ)", "ЩӮЫҢЩ…ШӘ ЩҲШұЩҲШҜ", "SL Ш§ЩҲЩ„ЫҢЩҮ",
               "Ш§ШіШӘШұШ§ШӘЪҳЫҢ", "ШЁШ§Щ„Ш§ШӘШұЫҢЩҶ R", "ЩҶШӘЫҢШ¬ЩҮ (R)", "ШӘШ§ШұЫҢШ® Ш®ШұЩҲШ¬ (Ш§ЫҢШұШ§ЩҶ)",
               "ЩӮЫҢЩ…ШӘ Ш®ШұЩҲШ¬", "ЩҲШ¶Ш№ЫҢШӘ"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        (_, _, _, symbol, direction, entry_time, entry_price, sl_price,
         strategy, max_r, result_r, exit_time, exit_price, status, _) = row
        ws.append([
            symbol, "Long" if direction == "long" else "Short",
            entry_time, entry_price, sl_price, strategy, max_r, result_r,
            exit_time or "---", exit_price or "---",
            "ШЁШіШӘЩҮ ШҙШҜЩҮ" if status == "closed" else "ШЁШ§ШІ"
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value)) for c in col if c.value is not None)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# в”Җв”Җв”Җ Щ…ЩҶЩҲЩҮШ§ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
def backtest_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("вһ• ШЁЪ©вҖҢШӘШіШӘ Ш¬ШҜЫҢШҜ", callback_data="bt_new")],
        [InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="back_financial")],
    ])

def strategy_choice_menu(user_id):
    strategies = get_strategies(user_id)
    keyboard = []
    for sid, name, _ in strategies[:10]:
        keyboard.append([InlineKeyboardButton(f"рҹ“Ӣ {name}", callback_data=f"bt_strat_use_{sid}")])
    keyboard.append([InlineKeyboardButton("рҹҶ• ШіШ§Ш®ШӘ Ш§ШіШӘШұШ§ШӘЪҳЫҢ Ш¬ШҜЫҢШҜ", callback_data="bt_strat_create")])
    keyboard.append([InlineKeyboardButton("рҹ”ҷ ШЁШұЪҜШҙШӘ", callback_data="fin_backtest")])
    return InlineKeyboardMarkup(keyboard)

def r_level_question_menu(r_mult):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("вң… ШЁЩ„ЩҮШҢ Ш¬Ш§ШЁШ¬Ш§ Ъ©ЩҶ", callback_data=f"bt_r_yes_{r_mult}")],
        [InlineKeyboardButton("вқҢ ЩҶЩҮШҢ ШұШҜ ШҙЩҲ", callback_data=f"bt_r_no_{r_mult}")],
    ])

def yes_no_menu(yes_cb, no_cb):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("вң… ШЁЩ„ЩҮ", callback_data=yes_cb)],
        [InlineKeyboardButton("вқҢ Ш®ЫҢШұ", callback_data=no_cb)],
    ])

def direction_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹ“Ҳ Long (Ш®ШұЫҢШҜ)", callback_data="bt_dir_long")],
        [InlineKeyboardButton("рҹ“ү Short (ЩҒШұЩҲШҙ)", callback_data="bt_dir_short")],
    ])

def timeframe_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("1 ШҜЩӮЫҢЩӮЩҮ (ШӘШ§ 7 ШұЩҲШІ ЩҫЫҢШҙ)", callback_data="bt_tf_1m")],
        [InlineKeyboardButton("15 ШҜЩӮЫҢЩӮЩҮ (ШӘШ§ 60 ШұЩҲШІ ЩҫЫҢШҙ)", callback_data="bt_tf_15m")],
        [InlineKeyboardButton("1 ШіШ§Ш№ШӘ (ШӘШ§ 2 ШіШ§Щ„ ЩҫЫҢШҙ)", callback_data="bt_tf_1h")],
    ])

# в”Җв”Җв”Җ ЩҮЩҶШҜЩ„Шұ ШҜЪ©Щ…ЩҮвҖҢЩҮШ§ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
async def handle_backtest(query, user_id):
    data = query.data

    if data == "fin_backtest":
        await query.edit_message_text(
            "рҹ“Ҡ ШЁЪ©вҖҢШӘШіШӘ Щ…ШҜЫҢШұЫҢШӘ ЩҫЩҲШІЫҢШҙЩҶ\n\n"
            "Ш§ЫҢЩҶ Ш§ШЁШІШ§Шұ ШЁШ§ ЪҜШұЩҒШӘЩҶ ЩҶЩӮШ·ЩҮ ЩҲШұЩҲШҜШҢ ШӯШҜ Ш¶ШұШұ ЩҲ Ш§ШіШӘШұШ§ШӘЪҳЫҢ ШҙШ®ШөЫҢ ШҙЩ…Ш§ШҢ "
            "Щ…ШіЫҢШұ ЩҲШ§ЩӮШ№ЫҢ ЩӮЫҢЩ…ШӘ ШұШ§ ШҙШЁЫҢЩҮвҖҢШіШ§ШІЫҢ Щ…ЫҢвҖҢЪ©ЩҶШҜ.\n\n"
            "ШіШ·ЩҲШӯ ШұЫҢЩҲШ§ШұШҜ ШЁШұШұШіЫҢвҖҢШҙШҜЩҮ:\n"
            "R0.5, R1, R1.5, R2, R3, R4, R8, R10\n"
            "(ШҜШұ R10 ЩҮЩ…ЫҢШҙЩҮ Ш®ШұЩҲШ¬ Ъ©Ш§Щ…Щ„)",
            reply_markup=backtest_menu()
        )

    elif data == "bt_new":
        set_state(user_id, "bt_mode_choice")
        await query.edit_message_text(
            "рҹ“Ҡ ШЁЪ©вҖҢШӘШіШӘ Ш¬ШҜЫҢШҜ\n\nЩҶЩҲШ№ ШЁЪ©вҖҢШӘШіШӘ ШұШ§ Ш§ЩҶШӘШ®Ш§ШЁ Ъ©ЩҶЫҢШҜ:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("рҹ”№ ШӘЪ©ЫҢ (ЫҢЪ© Ш§ШұШІ)", callback_data="bt_mode_single")],
                [InlineKeyboardButton("рҹ”ё ЪҜШұЩҲЩҮЫҢ (ЪҶЩҶШҜ Ш§ШұШІ)", callback_data="bt_mode_group")],
            ])
        )

    elif data in ["bt_mode_single", "bt_mode_group"]:
        mode = "single" if data == "bt_mode_single" else "group"
        set_state(user_id, "bt_strategy_choice", mode)
        strategies = get_strategies(user_id)
        if strategies:
            await query.edit_message_text(
                "рҹ“Ӣ Щ…ЫҢвҖҢШ®ЩҲШ§ЩҮЫҢШҜ Ш§ШІ Ш§ШіШӘШұШ§ШӘЪҳЫҢ ЩӮШЁЩ„ЫҢ Ш§ШіШӘЩҒШ§ШҜЩҮ Ъ©ЩҶЫҢШҜ ЫҢШ§ Ш¬ШҜЫҢШҜ ШЁШіШ§ШІЫҢШҜШҹ",
                reply_markup=strategy_choice_menu(user_id)
            )
        else:
            set_state(user_id, "bt_strat_r_0.5", f"{mode}|")
            await query.edit_message_text(
                "рҹҶ• ШіШ§Ш®ШӘ Ш§ШіШӘШұШ§ШӘЪҳЫҢ Ш¬ШҜЫҢШҜ\n\nШҜШұ ШіШ·Шӯ R0.5ШҢ ШӯШҜ Ш¶ШұШұ ШұШ§ Ш¬Ш§ШЁШ¬Ш§ Ъ©ЩҶЩ…Шҹ",
                reply_markup=r_level_question_menu("0.5")
            )

    elif data == "bt_strat_create":
        state, sdata = get_state(user_id)
        mode = sdata.split("|")[0] if sdata else "single"
        set_state(user_id, "bt_strat_r_0.5", f"{mode}|")
        await query.edit_message_text(
            "рҹҶ• ШіШ§Ш®ШӘ Ш§ШіШӘШұШ§ШӘЪҳЫҢ Ш¬ШҜЫҢШҜ\n\nШҜШұ ШіШ·Шӯ R0.5ШҢ ШӯШҜ Ш¶ШұШұ ШұШ§ Ш¬Ш§ШЁШ¬Ш§ Ъ©ЩҶЩ…Шҹ",
            reply_markup=r_level_question_menu("0.5")
        )

    elif data.startswith("bt_strat_use_"):
        sid = int(data.replace("bt_strat_use_", ""))
        name, rules = get_strategy_by_id(sid)
        state, sdata = get_state(user_id)
        mode = sdata.split("|")[0] if sdata else "single"
        if name:
            import json
            set_state(user_id, "bt_symbol", f"{mode}|{name}|{json.dumps(rules)}")
            await query.edit_message_text(
                f"вң… Ш§ШіШӘШұШ§ШӘЪҳЫҢ В«{name}В» Ш§ЩҶШӘШ®Ш§ШЁ ШҙШҜ.\n\n"
                f"рҹ“Ӣ Ш®Щ„Ш§ШөЩҮ ЩӮЩҲШ§ЩҶЫҢЩҶ Ш§ЫҢЩҶ Ш§ШіШӘШұШ§ШӘЪҳЫҢ:\n\n{describe_strategy(rules)}\n\n"
                f"ЩҶЩ…Ш§ШҜ Ш§ШұШІ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:\nЩ…Ш«Ш§Щ„: `EURUSD` `GOLD` `USDCAD`",
                parse_mode="Markdown"
            )

    elif data.startswith("bt_r_yes_") or data.startswith("bt_r_no_"):
        is_yes = data.startswith("bt_r_yes_")
        r_mult = data.replace("bt_r_yes_", "").replace("bt_r_no_", "")
        state, sdata = get_state(user_id)

        if is_yes:
            set_state(user_id, f"bt_strat_slval_{r_mult}", sdata)
            await query.edit_message_text(
                f"рҹҺҜ ШҜШұ ШіШ·Шӯ R{r_mult}ШҢ ШӯШҜ Ш¶ШұШұ ШұШ§ ШұЩҲЫҢ ЪҶЩҮ ШұЫҢЩҲШ§ШұШҜЫҢ ШЁЪҜШ°Ш§ШұЩ…Шҹ\n\n"
                f"ЫҢЪ© Ш№ШҜШҜ ШЁЫҢЩҶ `0` ЩҲ `{r_mult}` ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ.\n"
                f"Щ…Ш«Ш§Щ„: Ш§ЪҜШұ ШЁЩҶЩҲЫҢШіЫҢШҜ `0.2` ЫҢШ№ЩҶЫҢ ШӯШҜ Ш¶ШұШұ ШұЩҲЫҢ R0.2 ЩӮШұШ§Шұ Щ…ЫҢвҖҢЪҜЫҢШұШҜ.\n\n"
                f"ШЁШұШ§ЫҢ ШұЫҢШіЪ©вҖҢЩҒШұЫҢ ШҜЩӮЫҢЩӮ (ЩҶЩӮШ·ЩҮ ЩҲШұЩҲШҜ) ШЁЩҶЩҲЫҢШіЫҢШҜ `0`",
                parse_mode="Markdown"
            )
        else:
            mode, rules_str = (sdata.split("|", 1) + [""])[:2]
            import json
            rules = json.loads(rules_str) if rules_str else {}
            rules[r_mult] = None
            next_idx = R_LEVELS.index(float(r_mult)) + 1
            await proceed_to_next_r_level(query, user_id, mode, rules, next_idx)

# Ш№ШЁЩҲШұ ШЁЩҮ ШіШ·Шӯ R ШЁШ№ШҜЫҢ ЫҢШ§ ЩҫШ§ЫҢШ§ЩҶ ШіШ§Ш®ШӘ Ш§ШіШӘШұШ§ШӘЪҳЫҢ
async def proceed_to_next_r_level(query_or_update, user_id, mode, rules, next_idx, is_message=False):
    import json
    if next_idx < len(R_LEVELS) - 1:  # R10 ШіЩҲШ§Щ„ ЩҶШҜШ§ШұШҜ
        next_r = R_LEVELS[next_idx]
        set_state(user_id, f"bt_strat_r_{next_r}", f"{mode}|{json.dumps(rules)}")
        text = f"ШҜШұ ШіШ·Шӯ R{next_r}ШҢ ШӯШҜ Ш¶ШұШұ ШұШ§ Ш¬Ш§ШЁШ¬Ш§ Ъ©ЩҶЩ…Шҹ"
        markup = r_level_question_menu(str(next_r))
        if is_message:
            await query_or_update.message.reply_text(text, reply_markup=markup)
        else:
            await query_or_update.edit_message_text(text, reply_markup=markup)
    else:
        set_state(user_id, "bt_strat_name", f"{mode}|{json.dumps(rules)}")
        text = "вң… ШӘЩ…Ш§Щ… ШіШ·ЩҲШӯ ШӘЩҶШёЫҢЩ… ШҙШҜ!\n\nШ§ШіЩ…ЫҢ ШЁШұШ§ЫҢ Ш§ЫҢЩҶ Ш§ШіШӘШұШ§ШӘЪҳЫҢ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ (Щ…Ш«Ш§Щ„: Ш§ШіШӘШұШ§ШӘЪҳЫҢ Ш§ШөЩ„ЫҢ):"
        if is_message:
            await query_or_update.message.reply_text(text)
        else:
            await query_or_update.edit_message_text(text)

async def handle_backtest_message(user_id, text, update):
    state, data = get_state(user_id)
    text = text.strip()

    # в”Җв”Җв”Җ ШҜШұЫҢШ§ЩҒШӘ Щ…ЩӮШҜШ§Шұ SL ШЁШұШ§ЫҢ ЫҢЪ© ШіШ·Шӯ R Ш®Ш§Шө (ШЁЩҮвҖҢШөЩҲШұШӘ Ш№ШҜШҜ ШұЫҢЩҲШ§ШұШҜШҢ ЩҶЩҮ ЩӮЫҢЩ…ШӘ) в”Җв”Җв”Җ
    if state and state.startswith("bt_strat_slval_"):
        r_mult = state.replace("bt_strat_slval_", "")
        import json
        mode, rules_str = (data.split("|", 1) + [""])[:2]
        rules = json.loads(rules_str) if rules_str else {}

        try:
            r_value = float(text.replace(",", ""))
        except:
            await update.message.reply_text(
                f"вқҢ Щ„Ш·ЩҒШ§ЩӢ ЫҢЪ© Ш№ШҜШҜ Щ…Ш№ШӘШЁШұ ШЁЫҢЩҶ `0` ЩҲ `{r_mult}` ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:",
                parse_mode="Markdown"
            )
            return True

        current_r = float(r_mult)
        if r_value < 0 or r_value >= current_r:
            await update.message.reply_text(
                f"вқҢ Ш№ШҜШҜ ШЁШ§ЫҢШҜ ШЁЫҢЩҶ `0` ЩҲ `{r_mult}` ШЁШ§ШҙШҜ (Ъ©ЩҲЪҶЪ©вҖҢШӘШұ Ш§ШІ ШіШ·Шӯ ЩҒШ№Щ„ЫҢ).\n"
                f"Ш№ШҜШҜ ЩҲШ§ШұШҜвҖҢШҙШҜЩҮ ({r_value}) Ш®Ш§ШұШ¬ Ш§ШІ Ш§ЫҢЩҶ ШЁШ§ШІЩҮ Ш§ШіШӘ.\nШҜЩҲШЁШ§ШұЩҮ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:",
                parse_mode="Markdown"
            )
            return True

        # Ш°Ш®ЫҢШұЩҮ ШЁЩҮвҖҢШөЩҲШұШӘ Ш¶ШұЫҢШЁ ШұЫҢЩҲШ§ШұШҜ ЩҶШіШЁЫҢ (ЩҶЩҮ ЩӮЫҢЩ…ШӘ Щ…Ш·Щ„ЩӮ)
        # 0 ЫҢШ№ЩҶЫҢ ШҜЩӮЫҢЩӮШ§ЩӢ ЩҶЩӮШ·ЩҮ ЩҲШұЩҲШҜ (ШұЫҢШіЪ©вҖҢЩҒШұЫҢ Ъ©Ш§Щ…Щ„)
        rules[r_mult] = {"type": "r_trail", "value": r_value}

        next_idx = R_LEVELS.index(float(r_mult)) + 1
        await proceed_to_next_r_level(update, user_id, mode, rules, next_idx, is_message=True)
        return True

    # в”Җв”Җв”Җ Ш«ШЁШӘ ЩҶШ§Щ… Ш§ШіШӘШұШ§ШӘЪҳЫҢ в”Җв”Җв”Җ
    elif state == "bt_strat_name":
        import json
        mode, rules_str = (data.split("|", 1) + [""])[:2]
        rules = json.loads(rules_str) if rules_str else {}
        strategy_name = text
        save_strategy(user_id, strategy_name, rules)
        set_state(user_id, "bt_symbol", f"{mode}|{strategy_name}|{json.dumps(rules)}")
        await update.message.reply_text(
            f"вң… Ш§ШіШӘШұШ§ШӘЪҳЫҢ В«{strategy_name}В» Ш°Ш®ЫҢШұЩҮ ШҙШҜ.\n\n"
            f"рҹ“Ӣ Ш®Щ„Ш§ШөЩҮ ЩӮЩҲШ§ЩҶЫҢЩҶ Ш§ЫҢЩҶ Ш§ШіШӘШұШ§ШӘЪҳЫҢ:\n\n{describe_strategy(rules)}\n\n"
            f"ЩҶЩ…Ш§ШҜ Ш§ШұШІ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:\nЩ…Ш«Ш§Щ„: `EURUSD` `GOLD` `USDCAD`",
            parse_mode="Markdown"
        )
        return True

    # в”Җв”Җв”Җ ЩҶЩ…Ш§ШҜ в”Җв”Җв”Җ
    elif state == "bt_symbol":
        symbol = text.upper()
        parts = data.split("|")
        mode, strategy_name, rules_str = parts[0], parts[1], parts[2]
        # Ш§ЪҜШұ Ш§ЫҢЩҶ Ш§ШҜШ§Щ…ЩҮвҖҢЫҢ ЫҢЪ© batch ЪҜШұЩҲЩҮЫҢ Ш§ШіШӘШҢ batch_id ЩӮШЁЩ„ЫҢ ШұШ§ ШӯЩҒШё Ъ©ЩҶ
        if len(parts) > 3 and parts[3].startswith("__BATCH__"):
            batch_id = parts[3].replace("__BATCH__", "")
            set_state(user_id, "bt_direction", f"{mode}|{strategy_name}|{rules_str}|{symbol}|__BATCHID__{batch_id}")
        else:
            set_state(user_id, "bt_direction", f"{mode}|{strategy_name}|{rules_str}|{symbol}")
        await update.message.reply_text("Ш¬ЩҮШӘ Щ…Ш№Ш§Щ…Щ„ЩҮ ШұШ§ Ш§ЩҶШӘШ®Ш§ШЁ Ъ©ЩҶЫҢШҜ:", reply_markup=direction_menu())
        return True

    # в”Җв”Җв”Җ ШӘШ§ШұЫҢШ® ЩҲШұЩҲШҜ в”Җв”Җв”Җ
    elif state == "bt_entry_date":
        try:
            datetime.strptime(text, "%Y-%m-%d")
            set_state(user_id, "bt_entry_time", f"{data}|{text}")
            await update.message.reply_text(
                "вҸ° ШіШ§Ш№ШӘ ЩҲШұЩҲШҜ ШұШ§ ШЁЩҮ ЩҲЩӮШӘ Ш§ЫҢШұШ§ЩҶ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ (ЩҒШұЩ…ШӘ 24 ШіШ§Ш№ШӘЩҮ HH:MM)\nЩ…Ш«Ш§Щ„: `14:35`",
                parse_mode="Markdown"
            )
        except:
            await update.message.reply_text("вқҢ ЩҒШұЩ…ШӘ ШӘШ§ШұЫҢШ® Ш§ШҙШӘШЁШ§ЩҮ Ш§ШіШӘ.\nЩ…Ш«Ш§Щ„: `2026-03-10`", parse_mode="Markdown")
        return True

    # в”Җв”Җв”Җ ШіШ§Ш№ШӘ ЩҲШұЩҲШҜ в”Җв”Җв”Җ
    elif state == "bt_entry_time":
        try:
            datetime.strptime(text, "%H:%M")
            set_state(user_id, "bt_entry_price", f"{data}|{text}")
            await update.message.reply_text("ЩӮЫҢЩ…ШӘ ЩҲШұЩҲШҜ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        except:
            await update.message.reply_text("вқҢ ЩҒШұЩ…ШӘ ШіШ§Ш№ШӘ Ш§ШҙШӘШЁШ§ЩҮ Ш§ШіШӘ.\nЩ…Ш«Ш§Щ„: `14:35` (24 ШіШ§Ш№ШӘЩҮ)", parse_mode="Markdown")
        return True

    # в”Җв”Җв”Җ ЩӮЫҢЩ…ШӘ ЩҲШұЩҲШҜ в”Җв”Җв”Җ
    elif state == "bt_entry_price":
        try:
            entry_price = float(text.replace(",", ""))
            set_state(user_id, "bt_sl_price", f"{data}|{entry_price}")
            await update.message.reply_text("ЩӮЫҢЩ…ШӘ ШӯШҜ Ш¶ШұШұ Ш§ЩҲЩ„ЫҢЩҮ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        except:
            await update.message.reply_text("вқҢ Щ„Ш·ЩҒШ§ЩӢ ЫҢЪ© Ш№ШҜШҜ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        return True

    # в”Җв”Җв”Җ ШӯШҜ Ш¶ШұШұ Ш§ЩҲЩ„ЫҢЩҮ в”Җв”Җв”Җ
    elif state == "bt_sl_price":
        try:
            sl_price = float(text.replace(",", ""))
            parts = data.split("|")
            # Ш§ЪҜШұ Щ…ШіЫҢШұ ЪҜШұЩҲЩҮЫҢ Ш§ШіШӘШҢ ЫҢЪ© ЩӮШ·Ш№ЩҮ Ш§Ш¶Ш§ЩҒЩҮ (__BATCHID__xxx) ШҜШұ Щ…ЫҢШ§ЩҶЩҮ ЩҲШ¬ЩҲШҜ ШҜШ§ШұШҜ
            if len(parts) == 9 and parts[4].startswith("__BATCHID__"):
                mode, strategy_name, rules_str, symbol, batch_tag, direction, entry_date, entry_time, entry_price = parts
            else:
                mode, strategy_name, rules_str, symbol, direction, entry_date, entry_time, entry_price = parts
                batch_tag = None
            entry_price = float(entry_price)

            if (direction == "long" and sl_price >= entry_price) or (direction == "short" and sl_price <= entry_price):
                await update.message.reply_text("вқҢ ШӯШҜ Ш¶ШұШұ ШЁШ§ Ш¬ЩҮШӘ Щ…Ш№Ш§Щ…Щ„ЩҮ ЩҮЩ…Ш®ЩҲШ§ЩҶЫҢ ЩҶШҜШ§ШұШҜ.\nШҜЩҲШЁШ§ШұЩҮ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
                return True

            symbol_field = f"{symbol}{batch_tag}" if batch_tag else symbol
            set_state(user_id, "bt_timeframe",
                      f"{mode}|{strategy_name}|{rules_str}|{symbol_field}|{direction}|{entry_date}|{entry_time}|{entry_price}|{sl_price}")
            await update.message.reply_text("вҸұ ШӘШ§ЫҢЩ…вҖҢЩҒШұЫҢЩ… ШұШ§ Ш§ЩҶШӘШ®Ш§ШЁ Ъ©ЩҶЫҢШҜ:", reply_markup=timeframe_menu())
        except:
            await update.message.reply_text("вқҢ Щ„Ш·ЩҒШ§ЩӢ ЫҢЪ© Ш№ШҜШҜ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:")
        return True

    return False

# в”Җв”Җв”Җ ШҜЪ©Щ…ЩҮвҖҢЩҮШ§ЫҢ Ш¬ЩҮШӘ ЩҲ ШӘШ§ЫҢЩ…вҖҢЩҒШұЫҢЩ… ЩҲ Ш§ШҜШ§Щ…ЩҮ/ЩҫШ§ЫҢШ§ЩҶ ЪҜШұЩҲЩҮЫҢ в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ
async def handle_backtest_buttons_extra(query, user_id, data):
    if data in ["bt_dir_long", "bt_dir_short"]:
        direction = "long" if data == "bt_dir_long" else "short"
        state, sdata = get_state(user_id)
        if state == "bt_direction":
            set_state(user_id, "bt_entry_date", f"{sdata}|{direction}")
            await query.edit_message_text(
                f"вң… Ш¬ЩҮШӘ: {'Long рҹ“Ҳ' if direction=='long' else 'Short рҹ“ү'}\n\n"
                f"ШӘШ§ШұЫҢШ® ЩҲШұЩҲШҜ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ (ЩҒШұЩ…ШӘ: YYYY-MM-DD)\nЩ…Ш«Ш§Щ„: `2026-03-10`",
                parse_mode="Markdown"
            )
        return True

    if data.startswith("bt_tf_"):
        timeframe = data.replace("bt_tf_", "")
        state, sdata = get_state(user_id)
        if state == "bt_timeframe":
            set_state(user_id, "bt_processing", f"{sdata}|{timeframe}")
            await query.edit_message_text("вҸі ШҜШұ ШӯШ§Щ„ ШҜШұЫҢШ§ЩҒШӘ ШҜШ§ШҜЩҮвҖҢЩҮШ§ЫҢ ШӘШ§ШұЫҢШ®ЫҢ ЩҲ ШҙШЁЫҢЩҮвҖҢШіШ§ШІЫҢ...")
            await run_backtest_final(user_id, query)
        return True

    if data == "bt_add_another":
        state, sdata = get_state(user_id)
        parts = sdata.split("|")
        mode, strategy_name, rules_str, batch_id = parts[0], parts[1], parts[2], parts[3]
        set_state(user_id, "bt_symbol", f"{mode}|{strategy_name}|{rules_str}|__BATCH__{batch_id}")
        await query.edit_message_text(
            "вһ• ЩҶЩ…Ш§ШҜ Ш§ШұШІ ШЁШ№ШҜЫҢ ШұШ§ ЩҲШ§ШұШҜ Ъ©ЩҶЫҢШҜ:\nЩ…Ш«Ш§Щ„: `EURUSD` `GOLD` `USDCAD`",
            parse_mode="Markdown"
        )
        return True

    if data == "bt_finish_group":
        state, sdata = get_state(user_id)
        parts = sdata.split("|")
        batch_id = parts[3]
        await send_batch_excel(query, user_id, batch_id)
        clear_state(user_id)
        return True

    return False

async def run_backtest_final(user_id, query):
    state, sdata = get_state(user_id)
    if state != "bt_processing" or not sdata:
        return

    parts = sdata.split("|")
    mode, strategy_name, rules_str, symbol, direction, entry_date, entry_time, entry_price, sl_price, timeframe = parts
    entry_price, sl_price = float(entry_price), float(sl_price)

    import json
    rules = json.loads(rules_str)

    existing_batch_id = None
    if "__BATCHID__" in symbol:
        symbol, batch_tag = symbol.split("__BATCHID__")
        existing_batch_id = batch_tag

    batch_id = existing_batch_id if existing_batch_id else f"{user_id}_{int(datetime.now().timestamp())}"

    candles, entry_dt_utc, error = await get_historical_candles(symbol, entry_date, entry_time, timeframe)
    if error:
        await query.message.reply_text(error, reply_markup=backtest_menu())
        clear_state(user_id)
        return

    result = run_custom_simulation(candles, entry_price, sl_price, direction, rules)
    R = abs(entry_price - sl_price)

    entry_dt_iran = datetime.strptime(f"{entry_date} {entry_time}", "%Y-%m-%d %H:%M")
    exit_dt_iran = to_iran_time(result["exit_time_utc"]) if result["exit_time_utc"] else None

    # в”Җв”Җв”Җ ШӯШ§Щ„ШӘ ШҜЫҢШЁШ§ЪҜ Щ…ЩҲЩӮШӘ: ЩҶЩ…Ш§ЫҢШҙ ЪҶЩҶШҜ Ъ©ЩҶШҜЩ„ Ш§ЩҲЩ„ ШЁШұШ§ЫҢ ШЁШұШұШіЫҢ ШҜЩӮЫҢЩӮ ШҜШ§ШҜЩҮ Ш®Ш§Щ… в”Җв”Җв”Җ
    debug_lines = ["рҹ”Қ ШҜЫҢШЁШ§ЪҜ - Ъ©ЩҶШҜЩ„вҖҢЩҮШ§ЫҢ Ш§ШЁШӘШҜШ§ЫҢЫҢ (ЩҲЩӮШӘ Ш§ЫҢШұШ§ЩҶ):"]
    for c in candles[:8]:
        c_iran = to_iran_time(c["time_utc"])
        debug_lines.append(f"  {c_iran.strftime('%m-%d %H:%M')} | H={c['high']:.5f} L={c['low']:.5f}")
    debug_text = "\n".join(debug_lines)
    await query.message.reply_text(f"```\n{debug_text}\n```", parse_mode="Markdown")
    # в”Җв”Җв”Җ ЩҫШ§ЫҢШ§ЩҶ ШӯШ§Щ„ШӘ ШҜЫҢШЁШ§ЪҜ Щ…ЩҲЩӮШӘ в”Җв”Җв”Җ

    save_bt_result(
        user_id, batch_id, symbol, direction,
        entry_dt_iran.strftime("%Y-%m-%d %H:%M"), entry_price, sl_price, strategy_name,
        result["max_r"], result["result_r"],
        exit_dt_iran.strftime("%Y-%m-%d %H:%M") if exit_dt_iran else None,
        result["exit_price"], result["status"]
    )

    status_text = "рҹҹў ШЁШіШӘЩҮ ШҙШҜЩҮ" if result["status"] == "closed" else "рҹҹЎ ЩҮЩҶЩҲШІ ШЁШ§ШІ (ШӘШ§ ШўШ®ШұЫҢЩҶ ШҜШ§ШҜЩҮ)"
    exit_line = (f"Ш®ШұЩҲШ¬: {result['exit_price']:.5f}\nШӘШ§ШұЫҢШ® Ш®ШұЩҲШ¬ (Ш§ЫҢШұШ§ЩҶ): {exit_dt_iran.strftime('%Y-%m-%d %H:%M')}"
                 if result["exit_price"] else "ЩҮЩҶЩҲШІ Ш®Ш§ШұШ¬ ЩҶШҙШҜЩҮ")

    text = (
        f"рҹ“Ҡ ЩҶШӘЫҢШ¬ЩҮ ШЁЪ©вҖҢШӘШіШӘ вҖ” {symbol} ({'Long' if direction=='long' else 'Short'})\n"
        f"`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`\n"
        f"Ш§ШіШӘШұШ§ШӘЪҳЫҢ: {strategy_name}\n"
        f"ЩҲШұЩҲШҜ: {entry_price:.5f}\n"
        f"SL Ш§ЩҲЩ„ЫҢЩҮ: {sl_price:.5f}  (R = {R:.5f})\n"
        f"ШӘШ§ШұЫҢШ® ЩҲШұЩҲШҜ (Ш§ЫҢШұШ§ЩҶ): {entry_dt_iran.strftime('%Y-%m-%d %H:%M')}\n"
        f"ШӘШ§ЫҢЩ…вҖҢЩҒШұЫҢЩ…: {timeframe}\n"
        f"`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`\n"
        f"ЩҲШ¶Ш№ЫҢШӘ: {status_text}\n"
        f"ШЁШ§Щ„Ш§ШӘШұЫҢЩҶ R ШұШіЫҢШҜЩҮ: R{result['max_r']}\n"
        f"{exit_line}\n"
        f"ЩҶШӘЫҢШ¬ЩҮ: {'+' if result['result_r'] >= 0 else ''}{result['result_r']}R\n"
        f"`в”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғв”Ғ`"
    )

    if mode == "group":
        set_state(user_id, "bt_group_next", f"{mode}|{strategy_name}|{rules_str}|{batch_id}")
        await query.message.reply_text(
            text, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("вһ• Ш§ШұШІ ШЁШ№ШҜЫҢ", callback_data="bt_add_another")],
                [InlineKeyboardButton("вң… ЩҫШ§ЫҢШ§ЩҶ ЩҲ ШҜШұЫҢШ§ЩҒШӘ Ш§Ъ©ШіЩ„", callback_data="bt_finish_group")],
            ])
        )
    else:
        await send_single_excel(query, user_id, batch_id, text)
        clear_state(user_id)

async def send_single_excel(query, user_id, batch_id, summary_text):
    rows = get_batch_results(user_id, batch_id)
    excel_buffer = build_excel_report(rows)
    await query.message.reply_text(summary_text, parse_mode="Markdown", reply_markup=backtest_menu())
    await query.message.reply_document(
        document=excel_buffer, filename=f"backtest_{batch_id}.xlsx",
        caption="рҹ“Һ ЩҒШ§ЫҢЩ„ Ш§Ъ©ШіЩ„ ЩҶШӘЫҢШ¬ЩҮ ШЁЪ©вҖҢШӘШіШӘ"
    )

async def send_batch_excel(query, user_id, batch_id):
    rows = get_batch_results(user_id, batch_id)
    excel_buffer = build_excel_report(rows)

    lines = ["рҹ“Ҡ Ш®Щ„Ш§ШөЩҮ ЩҶШӘШ§ЫҢШ¬ ШЁЪ©вҖҢШӘШіШӘ ЪҜШұЩҲЩҮЫҢ:\n"]
    for row in rows:
        symbol, direction, max_r, result_r, status = row[3], row[4], row[9], row[10], row[13]
        emoji = "рҹҹў" if status == "closed" else "рҹҹЎ"
        lines.append(f"{emoji} {symbol} ({'Long' if direction=='long' else 'Short'}) вҶ’ R{max_r} | ЩҶШӘЫҢШ¬ЩҮ: {result_r:+.1f}R")

    await query.message.reply_text("\n".join(lines), reply_markup=backtest_menu())
    await query.message.reply_document(
        document=excel_buffer, filename=f"backtest_group_{batch_id}.xlsx",
        caption="рҹ“Һ ЩҒШ§ЫҢЩ„ Ш§Ъ©ШіЩ„ ЩҶШӘШ§ЫҢШ¬ ШЁЪ©вҖҢШӘШіШӘ ЪҜШұЩҲЩҮЫҢ"
    )

# в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ ЩҫШ§ЫҢШ§ЩҶ MODULE: BACKTEST в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җ



# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                      MAIN MENU                               в•‘
# в•‘                                                              в•‘
# в•‘  ШЁШұШ§ЫҢ Ш§Ш¶Ш§ЩҒЩҮ Ъ©ШұШҜЩҶ ШҜЪ©Щ…ЩҮ: ЫҢЩҮ Ш®Ш· InlineKeyboardButton Ш§Ш¶Ш§ЩҒЩҮ Ъ©ЩҶ в•‘
# в•‘  ШЁШұШ§ЫҢ ШӯШ°ЩҒ ШҜЪ©Щ…ЩҮ: Ш§ЩҲЩҶ Ш®Ш· ШұЩҲ ЩҫШ§Ъ© Ъ©ЩҶ                           в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ

def main_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("рҹ’° ШЁШ§ШІШ§ШұЩҮШ§ЫҢ Щ…Ш§Щ„ЫҢ",      callback_data="menu_financial")],   # вҶҗ ШӯШ°ЩҒ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
        [InlineKeyboardButton("рҹ“Ҡ ШӯШіШ§ШЁШҜШ§ШұЫҢ ШҙШ®ШөЫҢ",      callback_data="menu_accounting")],  # вҶҗ ШӯШ°ЩҒ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
        [InlineKeyboardButton("вҸ° ЫҢШ§ШҜШўЩҲШұ ЩҲ ШӘШіЪ©",        callback_data="menu_reminder")],    # вҶҗ ШӯШ°ЩҒ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
        [InlineKeyboardButton("рҹӨ– ШҜШіШӘЫҢШ§Шұ ЩҮЩҲШҙ Щ…ШөЩҶЩҲШ№ЫҢ",  callback_data="menu_ai")],          # вҶҗ ШӯШ°ЩҒ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
        [InlineKeyboardButton("рҹ‘¶ ЩҒШұШІЩҶШҜШ§ЩҶ",             callback_data="menu_kids")],        # вҶҗ ШӯШ°ЩҒ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
        [InlineKeyboardButton("рҹҺ® ШіШұЪҜШұЩ…ЫҢ",              callback_data="menu_fun")],         # вҶҗ ШӯШ°ЩҒ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
        [InlineKeyboardButton("рҹ“ҡ ЫҢШ§ШҜЪҜЫҢШұЫҢ",             callback_data="menu_learn")],       # вҶҗ ШӯШ°ЩҒ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
        [InlineKeyboardButton("вҡҷпёҸ ШӘЩҶШёЫҢЩ…Ш§ШӘ",             callback_data="menu_settings")],    # вҶҗ ШӯШ°ЩҒ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
        [InlineKeyboardButton("рҹҢҗ ШӘШұШ¬Щ…ЩҮ",  callback_data="menu_translate")],  # вҶҗ ШӯШ°ЩҒ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
    ])


# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                       ROUTER                                 в•‘
# в•‘                                                              в•‘
# в•‘  ЩҮШұ ШҜЪ©Щ…ЩҮ ШЁЩҮ Щ…Ш§ЪҳЩҲЩ„ Щ…ШұШЁЩҲШ·ЩҮ ЩҮШҜШ§ЫҢШӘ Щ…ЫҢШҙЩҮ                        в•‘
# в•‘  ШЁШұШ§ЫҢ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„: Ш®Ш· Щ…ШұШЁЩҲШ·ЩҮ ШұЩҲ ЩҫШ§Ъ© Ъ©ЩҶ                       в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    save_user(user.id, user.username or "", user.first_name)
    db_user = get_user(user.id)
    clear_state(user.id)

    if db_user[3] == 'active':
        expires = datetime.strptime(db_user[5], "%Y-%m-%d %H:%M")
        if datetime.now() >= expires:
            reject_user(user.id)
            await update.message.reply_text(f"вҸі {user.first_name} Ш№ШІЫҢШІШҢ\nШ§ШҙШӘШұШ§Ъ© ШҙЩ…Ш§ Щ…ЩҶЩӮШ¶ЫҢ ШҙШҜЩҮ.")
            await context.bot.send_message(ADMIN_ID, f"рҹ”„ Щ…ЩҶЩӮШ¶ЫҢ:\nрҹ‘Ө {user.first_name}\nрҹҶ” {user.id}")
            return
        await update.message.reply_text(f"вң… {user.first_name} Ш№ШІЫҢШІШҢ Ш®ЩҲШҙ ШўЩ…ШҜЫҢШҜ!\n\nрҹҸ  Щ…ЩҶЩҲЫҢ Ш§ШөЩ„ЫҢ:",
            reply_markup=main_menu())
        return

    if db_user[3] == 'rejected':
        await update.message.reply_text(f"вқҢ {user.first_name} Ш№ШІЫҢШІШҢ\nШҜШіШӘШұШіЫҢ ШҙЩ…Ш§ ШӘШЈЫҢЫҢШҜ ЩҶШҙШҜЩҮ.")
        return

    await update.message.reply_text(f"рҹ‘Ӣ ШіЩ„Ш§Щ… {user.first_name} Ш№ШІЫҢШІ!\n\nвң… ШҜШұШ®ЩҲШ§ШіШӘ Ш«ШЁШӘ ШҙШҜ.\nвҸі Щ…ЩҶШӘШёШұ ШӘШЈЫҢЫҢШҜ Ш§ШҜЩ…ЫҢЩҶ ШЁШ§ШҙЫҢШҜ.")
    await context.bot.send_message(ADMIN_ID,
        f"рҹ”” Ъ©Ш§ШұШЁШұ Ш¬ШҜЫҢШҜ:\nрҹ‘Ө {user.first_name}\nрҹҶ” {user.id}\nрҹ“ӣ @{user.username or 'ЩҶШҜШ§ШұШҜ'}",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("вң… ШӘШЈЫҢЫҢШҜ", callback_data=f"approve_{user.id}"),
            InlineKeyboardButton("вқҢ ШұШҜ", callback_data=f"reject_{user.id}")
        ]]))

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    user_id = query.from_user.id
    db_user = get_user(user_id)

    # в”Җв”Җв”Җ Ш§ШҜЩ…ЫҢЩҶ в”Җв”Җв”Җ
    if data.startswith("approve_") or data.startswith("reject_"):
        action, uid = data.split("_")
        uid = int(uid)
        target = get_user(uid)
        name = target[2] if target else str(uid)
        if action == "approve":
            approve_user(uid)
            target = get_user(uid)
            await query.edit_message_text(f"вң… {name} ШӘШЈЫҢЫҢШҜ ШҙШҜ.\nрҹ“… Ш§ЩҶЩӮШ¶Ш§: {target[5]}")
            await context.bot.send_message(uid, f"рҹҺү {name} Ш№ШІЫҢШІ!\nвң… ШӘШЈЫҢЫҢШҜ ШҙШҜЫҢШҜ.\nрҹ“… Ш§ШҙШӘШұШ§Ъ© ШӘШ§ {target[5]}\n\n/start ШЁШІЩҶЫҢШҜ.")
        else:
            reject_user(uid)
            await query.edit_message_text(f"вқҢ {name} ШұШҜ ШҙШҜ.")
            await context.bot.send_message(uid, f"вқҢ {name} Ш№ШІЫҢШІШҢ\nЩ…ШӘШЈШіЩҒШ§ЩҶЩҮ ШӘШЈЫҢЫҢШҜ ЩҶШҙШҜЫҢШҜ.")
        return

    if not db_user or db_user[3] != 'active':
        await query.answer("вқҢ ШҜШіШӘШұШіЫҢ ШҙЩ…Ш§ ЩҒШ№Ш§Щ„ ЩҶЫҢШіШӘ.", show_alert=True)
        return

    # в”Җв”Җв”Җ Щ…ЩҶЩҲЫҢ Ш§ШөЩ„ЫҢ в”Җв”Җв”Җ
    if data == "back_main":
        clear_state(user_id)
        await query.edit_message_text("рҹҸ  Щ…ЩҶЩҲЫҢ Ш§ШөЩ„ЫҢ:", reply_markup=main_menu())
        return

    # в”Җв”Җв”Җ ROUTER: ЩҮШұ prefix ШЁЩҮ Щ…Ш§ЪҳЩҲЩ„ Щ…ШұШЁЩҲШ·ЩҮ в”Җв”Җв”Җ
    if data.startswith("ps_"):
        await handle_position_size(query, user_id)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ Щ…ЫҢШІШ§ЩҶ Ш®ШұЫҢШҜ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif data.startswith("bt_") or data == "fin_backtest":
        handled = await handle_backtest_buttons_extra(query, user_id, data)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ ШЁЪ©вҖҢШӘШіШӘ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
        if not handled:
            await handle_backtest(query, user_id)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ ШЁЪ©вҖҢШӘШіШӘ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif data.startswith("menu_financial") or data.startswith("fin_") or data.startswith("wl_") or data.startswith("alarm_") or data.startswith("back_financial"):
        await handle_financial(query, user_id)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif data.startswith("menu_accounting") or data.startswith("acc_"):
        await handle_accounting(query, user_id)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif data.startswith("menu_reminder") or data.startswith("rem_"):
        await handle_reminder(query, user_id)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif data.startswith("menu_ai") or data.startswith("ai_"):
        await handle_ai(query, user_id)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif data.startswith("menu_kids") or data.startswith("kids_"):
        await handle_kids(query, user_id)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif (data.startswith("menu_fun") or data == "back_fun" or
          data.startswith("menu_health") or data.startswith("health_") or
          data.startswith("menu_book") or data.startswith("book_") or
          data.startswith("menu_music") or data.startswith("music_")):
        await handle_fun(query, user_id)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ ШіШұЪҜШұЩ…ЫҢ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif data.startswith("menu_learn") or data.startswith("learn_"):
        await handle_learning(query, user_id)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif data.startswith("menu_settings") or data.startswith("set_"):
        await handle_settings(query, user_id)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
    
    elif data.startswith("menu_translate") or data.startswith("tr_"):
        await handle_translate(query, user_id)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    else:
        await query.answer("рҹ”§ ШЁЩҮ ШІЩҲШҜЫҢ Ш§Ш¶Ш§ЩҒЩҮ Щ…ЫҢвҖҢШҙЩҲШҜ!", show_alert=True)
 
async def message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    db_user = get_user(user_id)
    if not db_user or db_user[3] != 'active':
        return

    text = update.message.text.strip()
    state, _ = get_state(user_id)


    if state in ["wl_add", "alarm_symbol", "alarm_price"]:
        await handle_financial_message(user_id, text, update)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif state in ["tr_waiting"]:
        await handle_translate_message(user_id, text, update)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif state and (state.startswith("book_") or state.startswith("music_")):
        await handle_fun_message(user_id, text, update)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ ШіШұЪҜШұЩ…ЫҢ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif state == "ai_chatting":
        await handle_ai_message(user_id, text, update)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ ШҜШіШӘЫҢШ§Шұ AI = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif state and state.startswith("kids_"):
        await handle_kids_message(user_id, text, update)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ ЩҒШұШІЩҶШҜШ§ЩҶ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif state and state.startswith("ps_"):
        await handle_position_size_message(user_id, text, update)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ Щ…ЫҢШІШ§ЩҶ Ш®ШұЫҢШҜ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif state and state.startswith("bt_"):
        await handle_backtest_message(user_id, text, update)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ ШЁЪ©вҖҢШӘШіШӘ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

    elif state and state.startswith("rem_"):
        await handle_reminder_message(user_id, text, update)  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ ЫҢШ§ШҜШўЩҲШұ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·

# в•”в•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•—
# в•‘                       RUN BOT                                в•‘
# в•ҡв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•җв•қ

async def post_init(app):
    asyncio.create_task(alarm_loop(app))
    asyncio.create_task(reminder_loop(app))  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ ЫҢШ§ШҜШўЩҲШұ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
    print("вҸ° ШіЫҢШіШӘЩ… ШўЩ„Ш§ШұЩ… ЩҒШ№Ш§Щ„ ШҙШҜ")

# в”Җв”Җв”Җ Web server ШіШ§ШҜЩҮ ШЁШұШ§ЫҢ Railway (ШЁШҜЩҲЩҶ Ш§ЫҢЩҶ Railway ШЁШ§ШӘ ШұЩҲ Ш®Ш§Щ…ЩҲШҙ Щ…ЫҢвҖҢЪ©ЩҶЩҮ) в”Җв”Җв”Җ
class HealthHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"Bot is running!")
    def log_message(self, *args):
        pass

def run_web():
    port = int(os.environ.get("PORT", 8080))
    server = HTTPServer(("0.0.0.0", port), HealthHandler)
    server.serve_forever()

Thread(target=run_web, daemon=True).start()

init_db()
init_backtest_db()  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ ШЁЪ©вҖҢШӘШіШӘ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
init_fun_db()  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ ШіШұЪҜШұЩ…ЫҢ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
init_ai_db()  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ ШҜШіШӘЫҢШ§Шұ AI = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
init_reminder_db()  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ ЫҢШ§ШҜШўЩҲШұ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
app = Application.builder().token(TOKEN).post_init(post_init).build()
app.add_handler(CommandHandler("start", start))
app.add_handler(CallbackQueryHandler(button_handler))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, message_handler))
app.add_handler(MessageHandler(filters.ChatType.CHANNEL, handle_channel_post))  # вҶҗ ШӯШ°ЩҒ Щ…Ш§ЪҳЩҲЩ„ ШіШұЪҜШұЩ…ЫҢ = ЩҫШ§Ъ© Ъ©ЩҶ Ш§ЫҢЩҶ Ш®Ш·
print("Bot is running...")
app.run_polling()